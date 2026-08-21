import pandas as pd
import streamlit as st
from datetime import datetime, timedelta
import gspread
from google.oauth2.service_account import Credentials
from config import USUARIOS
import uuid
import plotly.express as px
import os

from generar_hoja import generar_hoja_servicio
from supabase import create_client
supabase = create_client(
    st.secrets["SUPABASE_URL"],
    st.secrets["SUPABASE_KEY"]
)


NOMBRE_APP = "CRM Dashboard"
ICONO_APP = "🧹"

st.set_page_config(
    page_title=NOMBRE_APP,
    page_icon=ICONO_APP,
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    /* ── FONDO GENERAL ── */
    .stApp {
        background-color: #F5F7FA;
        color: #1A1D2E;
    }

    /* ── SIDEBAR ── */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1A1D2E 0%, #2D3561 100%);
        border-right: none;
    }
    [data-testid="stSidebar"] * { color: #F0F2F6 !important; }
    [data-testid="stSidebar"] > div:first-child { padding-top: 0rem !important; }
    section[data-testid="stSidebar"] > div { padding-top: 0.5rem !important; }

    /* ── MÉTRICAS ── */
    [data-testid="stMetric"] {
        background: #FFFFFF !important;
        border-radius: 12px;
        padding: 16px;
        border: 1px solid #E2E8F0;
        border-top: 3px solid #4F6AF0;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    }
    [data-testid="stMetricValue"] {
        font-size: clamp(16px, 2vw, 28px) !important;
        color: #1A1D2E !important;
        font-weight: 700 !important;
    }
    [data-testid="stMetricLabel"] {
        color: #64748B !important;
        font-size: 14px !important;
        font-weight: 500 !important;
    }

    /* ── TÍTULOS ── */
    h1 {
        color: #1A1D2E !important;
        font-weight: 700 !important;
        font-size: 2rem !important;
        letter-spacing: -0.5px;
    }
    h2, h3 {
        color: #4F6AF0 !important;
        font-weight: 600 !important;
        font-size: 1.2rem !important;
    }

    /* ── TEXTO GENERAL ── */
    p, span, label {
        color: #1A1D2E !important;
        font-size: 15px !important;
    }

    /* ── LABELS DE INPUTS ── */
    .stSelectbox label,
    .stTextInput label,
    .stNumberInput label,
    .stDateInput label,
    .stSlider label,
    .stCheckbox label,
    .stTextArea label,
    .stTimeInput label {
        color: #1A1D2E !important;
        font-size: 15px !important;
        font-weight: 500 !important;
    }

    /* ── INPUTS ── */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stTextArea > div > div > textarea {
        background-color: #FFFFFF !important;
        border: 1px solid #CBD5E1 !important;
        border-radius: 8px !important;
        color: #1A1D2E !important;
        font-size: 15px !important;
    }
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus {
        border: 1px solid #4F6AF0 !important;
        box-shadow: 0 0 0 3px rgba(79,106,240,0.1) !important;
    }

    /* ── SELECTBOX ── */
    .stSelectbox > div > div {
        background-color: #FFFFFF !important;
        border: 1px solid #CBD5E1 !important;
        border-radius: 8px !important;
        color: #1A1D2E !important;
        font-size: 15px !important;
    }

    /* ── DROPDOWN ABIERTO ── */
    [data-baseweb="popover"],
    [data-baseweb="menu"],
    [data-baseweb="option"] {
        background-color: #FFFFFF !important;
        color: #1A1D2E !important;
        font-size: 15px !important;
    }
    [data-baseweb="option"]:hover {
        background-color: #EEF2FF !important;
        color: #1A1D2E !important;
    }

    /* ── DATE PICKER ── */
    [data-baseweb="calendar"] {
        background-color: #FFFFFF !important;
        border-radius: 8px !important;
        border: 1px solid #CBD5E1 !important;
    }
    [data-baseweb="calendar"] * {
        color: #1A1D2E !important;
        background-color: #FFFFFF !important;
    }
    [data-baseweb="calendar"] button:hover {
        background-color: #EEF2FF !important;
    }
    [data-baseweb="calendar"] [aria-selected="true"],
    [data-baseweb="calendar"] [aria-selected="true"] * {
        background-color: #4F6AF0 !important;
        color: #FFFFFF !important;
    }
    [data-baseweb="calendar"] [aria-disabled="true"],
    [data-baseweb="calendar"] [aria-disabled="true"] * {
        color: #CBD5E1 !important;
        background-color: #FFFFFF !important;
    }

    /* ── BOTONES NORMALES ── */
    .stButton > button {
        background: linear-gradient(135deg, #4F6AF0 0%, #7C3AED 100%);
        color: white !important;
        border: none;
        border-radius: 8px;
        padding: 10px 24px;
        font-size: 15px;
        width: 100%;
        font-weight: 600;
        box-shadow: 0 2px 4px rgba(79,106,240,0.3);
    }
    .stButton > button:hover {
        opacity: 0.9;
        box-shadow: 0 4px 8px rgba(79,106,240,0.4);
    }

    /* ── FORM SUBMIT BUTTON ── */
    [data-testid="stFormSubmitButton"] > button {
        background: linear-gradient(135deg, #4F6AF0 0%, #7C3AED 100%) !important;
        color: white !important;
        border: none !important;
        font-weight: 600 !important;
        font-size: 16px !important;
        border-radius: 8px !important;
        width: 100% !important;
        padding: 12px 24px !important;
        box-shadow: 0 2px 4px rgba(79,106,240,0.3) !important;
    }

    /* ── FORM CONTAINER ── */
    .stForm {
        background-color: #FFFFFF !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 12px !important;
        padding: 24px !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06) !important;
    }

    /* ── DATAFRAME ── */
    .stDataFrame {
        border: 1px solid #E2E8F0;
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    }

    /* ── CAPTION ── */
    .stCaption {
        color: #64748B !important;
        font-size: 13px !important;
    }

    /* ── DIVISORES ── */
    hr { border-color: #E2E8F0 !important; }

    /* ── EXPANDER ── */
    .streamlit-expanderHeader {
        background-color: #FFFFFF !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 8px !important;
        color: #1A1D2E !important;
        font-size: 15px !important;
    }

    /* ── ALERTS ── */
    .stSuccess {
        background-color: #F0FDF4 !important;
        border: 1px solid #86EFAC !important;
        border-radius: 8px !important;
        color: #166534 !important;
    }
    .stError {
        background-color: #FEF2F2 !important;
        border: 1px solid #FCA5A5 !important;
        border-radius: 8px !important;
        color: #991B1B !important;
    }
    .stWarning {
        background-color: #FFFBEB !important;
        border: 1px solid #FCD34D !important;
        border-radius: 8px !important;
        color: #92400E !important;
    }
    .stInfo {
        background-color: #EEF2FF !important;
        border: 1px solid #A5B4FC !important;
        border-radius: 8px !important;
        color: #3730A3 !important;
    }

    /* ── MARKDOWN BOLD ── */
    strong { color: #1A1D2E !important; }

    /* ── PLACEHOLDER ── */
    input::placeholder {
        color: #94A3B8 !important;
        opacity: 1 !important;
    }

    /* ── SIDEBAR BOTONES ── */
    [data-testid="stSidebar"] .stButton > button {
        background: rgba(255,255,255,0.1) !important;
        border: 1px solid rgba(255,255,255,0.2) !important;
        color: #F0F2F6 !important;
        box-shadow: none !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(255,255,255,0.2) !important;
    }
</style>
""", unsafe_allow_html=True)

@st.cache_resource
def get_gspread_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",  # ← quita el .readonly
        "https://www.googleapis.com/auth/drive"           # ← quita el .readonly
    ]
    creds = Credentials.from_service_account_info(
        dict(st.secrets["google_credentials"]),
        scopes=scopes
    )
    return gspread.authorize(creds)
import time
def asignar_ids_clientes():
    import unicodedata

    client = get_gspread_client()
    sheet_ids = st.session_state.get("SHEET_IDS", {})

    def normalizar(nombre, tel=""):
        nombre = str(nombre).strip().lower()
        nombre = unicodedata.normalize("NFKD", nombre)
        nombre = "".join(c for c in nombre if not unicodedata.combining(c))
        nombre = " ".join(nombre.split())
        # Incluir los últimos 4 dígitos del tel para distinguir homónimos
        tel = str(tel).strip().replace("-", "").replace(" ", "")
        sufijo = tel[-4:] if len(tel) >= 4 else tel
        return f"{nombre}_{sufijo}"

    datos_sheets = {}
    for año, sheet_id in sheet_ids.items():
        if not sheet_id:
            continue
        try:
            sh = client.open_by_key(sheet_id)
            ws = sh.get_worksheet(0)

            # ── Verificar si ya existe columna ID Cliente ──
            headers = ws.row_values(1)
            if "ID Cliente" in headers:
                st.info(f"Sheet {año} ya tiene columna ID Cliente — se actualizarán los IDs sin insertar columna nueva.")
                col_id = headers.index("ID Cliente") + 1  # base 1
                col_nombre = headers.index("Nombre") + 1 if "Nombre" in headers else None
                col_tel = headers.index("Tel") + 1 if "Tel" in headers else None
                datos_sheets[año] = {
                    "ws": ws, "sh": sh,
                    "ya_tiene_columna": True,
                    "col_id": col_id,
                    "col_nombre": col_nombre,
                    "col_tel": col_tel,
                    "total_filas": len(ws.col_values(col_nombre)) if col_nombre else 0
                }
            else:
                col_nombre = headers.index("Nombre") + 1 if "Nombre" in headers else 4
                col_tel = headers.index("Tel") + 1 if "Tel" in headers else 5
                col_nombres = ws.col_values(col_nombre)
                col_tels = ws.col_values(col_tel)
                datos_sheets[año] = {
                    "ws": ws, "sh": sh,
                    "ya_tiene_columna": False,
                    "col_nombre": col_nombre,
                    "col_tel": col_tel,
                    "nombres": col_nombres,
                    "tels": col_tels
                }
        except Exception as e:
            st.warning(f"No se pudo leer el sheet {año}: {e}")

    if not datos_sheets:
        st.error("No hay sheets disponibles.")
        return

    # ── Construir mapa nombre+tel → ID ──
    mapa_id = {}
    contador = 1

    for año, data in datos_sheets.items():
        ws = data["ws"]
        if data["ya_tiene_columna"]:
            col_nombre = data["col_nombre"]
            col_tel = data["col_tel"]
            nombres = ws.col_values(col_nombre)[1:] if col_nombre else []
            tels = ws.col_values(col_tel)[1:] if col_tel else []
        else:
            nombres = data["nombres"][1:]
            tels = data["tels"][1:] if len(data.get("tels", [])) > 1 else [""] * len(nombres)

        for i, nombre in enumerate(nombres):
            if not nombre or str(nombre).strip() in ["", "nan"]:
                continue
            tel = tels[i] if i < len(tels) else ""
            clave = normalizar(nombre, tel)
            if clave not in mapa_id:
                mapa_id[clave] = contador
                contador += 1

    st.info(f"Se identificaron {contador - 1} clientes únicos (por nombre + teléfono).")

    # ── Insertar columna o actualizar IDs ──
    for año, data in datos_sheets.items():
        ws = data["ws"]
        sh = data["sh"]

        try:
            if not data["ya_tiene_columna"]:
                # Insertar columna D nueva
                sheet_tab_id = ws.id
                sh.batch_update({"requests": [{
                    "insertDimension": {
                        "range": {
                            "sheetId": sheet_tab_id,
                            "dimension": "COLUMNS",
                            "startIndex": 3,
                            "endIndex": 4
                        },
                        "inheritFromBefore": False
                    }
                }]})
                ws.update_cell(1, 4, "ID Cliente")
                col_id = 4
                col_nombre = data["col_nombre"] + 1  # se desplazó
                col_tel = data["col_tel"] + 1
            else:
                col_id = data["col_id"]
                col_nombre = data["col_nombre"]
                col_tel = data["col_tel"]

            # Leer nombres y tels actuales
            nombres = ws.col_values(col_nombre)[1:]
            tels = ws.col_values(col_tel)[1:] if col_tel else []

            updates = []
            for i, nombre in enumerate(nombres):
                if not nombre or str(nombre).strip() in ["", "nan"]:
                    continue
                tel = tels[i] if i < len(tels) else ""
                clave = normalizar(nombre, tel)
                id_cliente = mapa_id.get(clave, "")
                updates.append({
                    "range": f"{chr(64 + col_id)}{i + 2}",
                    "values": [[id_cliente]]
                })

            for i in range(0, len(updates), 100):
                ws.batch_update(updates[i:i+100])

            st.success(f"✅ Sheet {año} — {len(updates)} filas actualizadas")

        except Exception as e:
            st.error(f"Error en sheet {año}: {e}")

    st.success("🎉 Listo — IDs únicos asignados por nombre + teléfono")
    st.cache_data.clear()

# 🔥 FUNCIÓN AGREGAR CLIENTES (igual pero mejorada leve)
def agregar_a_sheets(data):
    client = get_gspread_client()
    sheet_id = SHEET_IDS[data["Año"].iloc[0]]
    sh = client.open_by_key(sheet_id)
    worksheet = sh.get_worksheet(0)

    # 🔥 folio único
    folio = str(int(datetime.now().timestamp()))

    worksheet.append_row([
        folio,                      # Folio sistema
        "",                         # Folio interno
        data["Fecha"].iloc[0],      # Fecha
        data["Nombre"].iloc[0],     # Nombre
        data["Tel"].iloc[0],        # Tel
        "",                         # Dirección
        data["Origen"].iloc[0],     # Origen
        data["Monto"].iloc[0],      # Monto
        data["Servicio"].iloc[0],   # Servicio
        "",                         # Comentarios
        "", "", ""                  # 90 días, 6 meses, 1 año
    ])

# ── LOGIN ──
def login():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(
            "<h1 style='text-align:center; color:#2B5BAA; font-size:2.5rem;'>CRM</h1>",
            unsafe_allow_html=True
        )
        st.markdown(
            "<h3 style='text-align:center; color:#2B5BAA;'>Iniciar sesión</h3>",
            unsafe_allow_html=True
        )
        st.markdown("<br>", unsafe_allow_html=True)

        with st.form("login_form"):
            email = st.text_input("Email", placeholder="Tu email")
            password = st.text_input(
                "Contraseña",
                type="password",
                placeholder="Tu contraseña"
            )
            submitted = st.form_submit_button(
                "Entrar",
                use_container_width=True
            )

            if submitted:
                try:
                    response = supabase.auth.sign_in_with_password({
                        "email": email,
                        "password": password
                    })

                    user_data = supabase.table("usuarios")\
                        .select("*")\
                        .eq("auth_id", response.user.id)\
                        .single()\
                        .execute()

                    st.session_state["usuario"] = user_data.data["username"]
                    st.session_state["empresa"] = user_data.data["empresa"]
                    st.session_state["sistema"] = user_data.data["sistema"]
                    st.session_state["auth_id"] = str(response.user.id)
                    st.session_state["empresa_id"] = user_data.data["id"]
                    st.session_state["access_token"] = response.session.access_token
                    st.session_state["refresh_token"] = response.session.refresh_token
                    st.rerun()

                except Exception as e:
                    st.error("Email o contraseña incorrectos")


if "usuario" not in st.session_state:
    login()
    st.stop()

if "cache_limpiado" not in st.session_state:
    st.cache_data.clear()
    st.session_state["cache_limpiado"] = True
# ── CONFIG DINÁMICA ──
NOMBRE_APP = USUARIOS.get(st.session_state["usuario"], {}).get("app_nombre", "CRM Dashboard")

# ── CARGAR CONFIG DESDE SUPABASE ──
@st.cache_data(ttl=601)
def cargar_config(empresa_id, access_token):
    try:
        client = create_client(
            st.secrets["SUPABASE_URL"],
            st.secrets["SUPABASE_KEY"]
        )
        client.postgrest.auth(access_token)

        sheets_resp = client.table("usuario_sheets")\
            .select("año, sheet_id")\
            .eq("empresa_id", empresa_id)\
            .execute()
        sheets = {int(r["año"]): r["sheet_id"] for r in sheets_resp.data} if sheets_resp.data else {}

        finanzas_resp = client.table("usuario_finanzas")\
            .select("año, url")\
            .eq("empresa_id", empresa_id)\
            .execute()
        finanzas = {int(r["año"]): r["url"] for r in finanzas_resp.data} if finanzas_resp.data else {}

        origenes_resp = client.table("usuario_origenes")\
            .select("clave, valor")\
            .eq("empresa_id", empresa_id)\
            .execute()
        origenes = {r["clave"]: r["valor"] for r in origenes_resp.data} if origenes_resp.data else {}

        plantillas_resp = client.table("usuario_plantillas")\
            .select("clave, mensaje")\
            .eq("empresa_id", empresa_id)\
            .execute()
        plantillas = {r["clave"]: r["mensaje"] for r in plantillas_resp.data} if plantillas_resp.data else {}

        categorias_resp = client.table("usuario_categorias")\
            .select("categoria, keywords")\
            .eq("empresa_id", empresa_id)\
            .execute()
        categorias = {r["categoria"]: r["keywords"] for r in categorias_resp.data} if categorias_resp.data else {}

        cotizador_resp = client.table("usuario_cotizador")\
            .select("*")\
            .eq("empresa_id", empresa_id)\
            .limit(1)\
            .execute()
        cotizador = {}
        if cotizador_resp.data:
            c = cotizador_resp.data[0]
            cotizador = {
                "paquetes": c.get("paquetes", []),
                "minimo": c.get("minimo", 950),
                "intro": c.get("intro", ""),
                "purt_descripcion": c.get("purt_descripcion", ""),
                "purt_costo": c.get("purt_costo", 0),
                "cierre": c.get("cierre", ""),
                "firma": c.get("firma", ""),
                "servicios_cantidad": c.get("servicios_cantidad", []),
                "servicios_plazas": c.get("servicios_plazas", []),
                "servicios_sillas": c.get("servicios_sillas", []),
            }

        precios_resp = client.table("usuario_cotizador_precios")\
            .select("servicio, paquete, precio")\
            .eq("empresa_id", empresa_id)\
            .execute()
        precios = {}
        if precios_resp.data:
            for r in precios_resp.data:
                if r["servicio"] not in precios:
                    precios[r["servicio"]] = {}
                precios[r["servicio"]][r["paquete"]] = float(r["precio"])
        if cotizador:
            cotizador["precios"] = precios

        columnas_resp = client.table("usuario_sheet_columnas")\
            .select("*")\
            .eq("empresa_id", empresa_id)\
            .limit(1)\
            .execute()
        sheet_columnas = {}
        if columnas_resp.data:
            c = columnas_resp.data[0]
            sheet_columnas = {
                "folio_sistema": c.get("col_folio_sistema", 1) - 1,
                "folio_interno": c.get("col_folio_interno", 2) - 1,
                "fecha": c.get("col_fecha", 3) - 1,
                "id_cliente": c.get("col_id_cliente", 4) - 1,
                "nombre": c.get("col_nombre", 5) - 1,
                "tel": c.get("col_tel", 6) - 1,
                "direccion": c.get("col_direccion", 7) - 1,
                "origen": c.get("col_origen", 8) - 1,
                "monto": c.get("col_monto", 9) - 1,
                "servicio": c.get("col_servicio", 10) - 1,
                "comentarios": c.get("col_comentarios", 11) - 1,
            }

        usuario_resp = client.table("usuarios")\
            .select("template_pdf, logo_url, app_nombre, app_icono, ciudad, estadisticas_hoja, features")\
            .eq("id", empresa_id)\
            .single()\
            .execute()
        template_pdf = usuario_resp.data.get("template_pdf") if usuario_resp.data else None
        logo_url = usuario_resp.data.get("logo_url") if usuario_resp.data else None
        app_nombre = usuario_resp.data.get("app_nombre", "CRM Dashboard") if usuario_resp.data else "CRM Dashboard"
        app_icono = usuario_resp.data.get("app_icono", "📊") if usuario_resp.data else "📊"
        ciudad = usuario_resp.data.get("ciudad", "") if usuario_resp.data else ""
        estadisticas_hoja = usuario_resp.data.get("estadisticas_hoja", "Estadisticas finales") if usuario_resp.data else "Estadisticas finales"
        features = usuario_resp.data.get("features") if usuario_resp.data else {}
        if not isinstance(features, dict):
            features = {}

        return {
            "sheets": sheets,
            "finanzas": finanzas,
            "origenes": origenes,
            "plantillas": plantillas,
            "categorias": categorias,
            "cotizador": cotizador,
            "template_pdf": template_pdf,
            "logo_url": logo_url,
            "app_nombre": app_nombre,
            "app_icono": app_icono,
            "ciudad": ciudad,
            "sheet_columnas": sheet_columnas,
            "estadisticas_hoja": estadisticas_hoja,
            "features": features,
        }

    except Exception as e:
        st.error(f"Error cargando config: {e}")
        return {}

empresa_id_config = st.session_state.get("empresa_id", "")
access_token_config = st.session_state.get("access_token", "")
config_usuario = cargar_config(empresa_id_config, access_token_config)

if st.session_state.get("usuario") and config_usuario:
    if st.session_state["usuario"] not in USUARIOS:
        USUARIOS[st.session_state["usuario"]] = {}
    USUARIOS[st.session_state["usuario"]]["sheets"] = config_usuario.get("sheets", {})
    USUARIOS[st.session_state["usuario"]]["finanzas"] = config_usuario.get("finanzas", {})
    USUARIOS[st.session_state["usuario"]]["origenes"] = config_usuario.get("origenes", {})
    USUARIOS[st.session_state["usuario"]]["plantillas"] = config_usuario.get("plantillas", {})
    USUARIOS[st.session_state["usuario"]]["categorias"] = config_usuario.get("categorias", {})
    USUARIOS[st.session_state["usuario"]]["cotizador"] = config_usuario.get("cotizador", {})
    USUARIOS[st.session_state["usuario"]]["template_pdf"] = config_usuario.get("template_pdf")
    USUARIOS[st.session_state["usuario"]]["logo_url"] = config_usuario.get("logo_url")
    USUARIOS[st.session_state["usuario"]]["app_nombre"] = config_usuario.get("app_nombre", "CRM Dashboard")
    USUARIOS[st.session_state["usuario"]]["app_icono"] = config_usuario.get("app_icono", "📊")
    USUARIOS[st.session_state["usuario"]]["ciudad"] = config_usuario.get("ciudad", "")
    USUARIOS[st.session_state["usuario"]]["sheet_columnas"] = config_usuario.get("sheet_columnas", {})
    USUARIOS[st.session_state["usuario"]]["estadisticas_hoja"] = config_usuario.get("estadisticas_hoja", "Estadisticas finales")
    USUARIOS[st.session_state["usuario"]]["features"] = config_usuario.get("features", {})

NOMBRE_APP = USUARIOS.get(st.session_state["usuario"], {}).get("app_nombre", "CRM Dashboard")
# ── CARGAR DATOS DESDE SUPABASE ──
@st.cache_data(ttl=300)
def cargar_datos(empresa_id, access_token):
    try:
        client = create_client(
            st.secrets["SUPABASE_URL"],
            st.secrets["SUPABASE_KEY"]
        )
        client.postgrest.auth(access_token)

        todos = []
        page = 0
        page_size = 500

        while True:
            response = client.table("clientes")\
                .select("*")\
                .eq("empresa_id", empresa_id)\
                .range(page * page_size, (page + 1) * page_size - 1)\
                .execute()

            if not response.data:
                break

            todos.extend(response.data)

            if len(response.data) < page_size:
                break

            page += 1

        if not todos:
            return pd.DataFrame()

        df = pd.DataFrame(todos)
        df = df.rename(columns={
            "comentarios": "Comentarios con llamada posterior a venta",
            "nombre": "Nombre",
            "tel": "Tel",
            "direccion": "Dirección",
            "origen": "Origen",
            "monto": "Monto",
            "servicio": "Servicio",
            "fecha": "Fecha",
            "año": "Año",
            "cliente_id": "ID Cliente"
        })
        return df

    except Exception as e:
        st.error(f"Error cargando datos: {e}")
        return pd.DataFrame()

empresa_id = st.session_state.get("empresa_id", "")
access_token = st.session_state.get("access_token", "")
df = cargar_datos(empresa_id, access_token)

if df is None or df.empty:
    df = pd.DataFrame({
        "Nombre": [], "Tel": [], "Fecha": [], "Monto": [],
        "Servicio": [], "Origen": [],
        "Comentarios con llamada posterior a venta": [], "Año": []
    })

df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
df["Monto"] = pd.to_numeric(df["Monto"], errors="coerce")
df["Mes"] = df["Fecha"].dt.month
df["Año"] = pd.to_numeric(df["Año"], errors="coerce").astype("Int64")

años_disponibles = sorted(df["Año"].dropna().unique().tolist())
if not años_disponibles:
    años_disponibles = [datetime.now().year]
años_sin_2026 = años_disponibles
# ── SIDEBAR ──
with st.sidebar:
    logo_url = USUARIOS[st.session_state["usuario"]].get("logo_url")
    if logo_url:
        st.image(logo_url, width=120)
    st.markdown(f"<h3 style='color:white'>{st.session_state['empresa']}</h3>", unsafe_allow_html=True)
    st.markdown("---")
    paginas = ["Resumen", "Ventas", "Clientes", "Servicios", "Follow Up", "Agenda", "Cotizaciones", "Chat", "Mi cuenta"]

    if "pagina" not in st.session_state:
        st.session_state["pagina"] = "Resumen"

    for p in paginas:
        if st.button(p, key=f"sidebar_{p}", use_container_width=True):
            st.session_state["pagina"] = p

    st.markdown("---")
    st.caption("Datos actualizados cada 5 min")

    if st.button("Actualizar datos", key="sidebar_actualizar", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    st.markdown("---")

    if st.button("Cerrar sesión", key="sidebar_logout", use_container_width=True):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()
    # ── RESUMEN ──
    # 🔥 IMPORTANTE: ESTO VA FUERA DEL SIDEBAR
pagina = st.session_state["pagina"]
def limpiar_numero(valor):
    if pd.isna(valor):
        return 0

    valor = str(valor)
    valor = valor.replace("$", "")
    valor = valor.replace(",", "")
    valor = valor.replace(" ", "")

    if valor == "-" or valor == "":
        return 0

    try:
        return float(valor)
    except:
        return 0

def cargar_finanzas(url):
    try:
        df = pd.read_csv(url, header=None)
    except Exception as e:
        st.error(f"Error descargando CSV: {e}")
        return None, None, None

    def sumar_columna_desde_fila(keyword, col_idx, filas_max=50):
        # Encontrar fila del keyword
        fila_inicio = None
        for i, row in df.iterrows():
            if str(row.iloc[0]).strip() == keyword:
                fila_inicio = i
                break
        if fila_inicio is None:
            return 0

        # Buscar fila del siguiente keyword para saber dónde parar
        fila_fin = fila_inicio + filas_max

        total = 0
        for i in range(fila_inicio, min(fila_fin, len(df))):
            try:
                val = df.iloc[i, col_idx]
                limpio = str(val).replace("$", "").replace(",", "").replace("`", "").strip()
                num = float(limpio)
                if num > 0:
                    total += num
            except:
                continue
        return total

    # Detectar si es estructura semanal (columna 42 = Total Mes)
    es_semanal = df.iloc[8, 42] == "Total Mes" if len(df.columns) > 42 else False

    if es_semanal:
        # Sumar columna 42 entre ENTRADAS y SALIDAS
        ingresos = sumar_columna_desde_fila("ENTRADAS", 42, filas_max=15)
        gastos = sumar_columna_desde_fila("SALIDAS", 42, filas_max=15)
    else:
        # Estructura mensual normal — buscar max en fila de Total Entradas/Salidas
        def buscar_total(keyword):
            filas = df[df.astype(str).apply(
                lambda row: row.str.contains(keyword, case=False, na=False).any(),
                axis=1
            )]
            if filas.empty:
                return 0
            fila = filas.iloc[0]
            valores = []
            for v in fila:
                try:
                    limpio = str(v).replace("$", "").replace(",", "").replace("`", "").strip()
                    num = float(limpio)
                    if num > 0:
                        valores.append(num)
                except:
                    continue
            return max(valores) if valores else 0

        ingresos = buscar_total("Total Entradas")
        gastos = buscar_total("Total Salidas")

    utilidad = ingresos - gastos
    return ingresos, gastos, utilidad
# ── RESUMEN ──
if pagina == "Resumen":
    st.title(NOMBRE_APP)

    # ─────────────────────────────
    # 🔔 BANNER DE RECORDATORIOS
    # ─────────────────────────────
    if "meses_recordatorio" not in st.session_state:
        st.session_state["meses_recordatorio"] = 6

    meses_recordatorio = st.session_state["meses_recordatorio"]

    df_banner = df.copy()
    df_banner["Fecha"] = pd.to_datetime(df_banner["Fecha"], errors="coerce")

    ultimo_banner = df_banner.groupby("Nombre").agg(
        Ultima_Visita=("Fecha", "max")
    ).reset_index()
    ultimo_banner["Meses_sin_servicio"] = (
        (datetime.now() - ultimo_banner["Ultima_Visita"]).dt.days / 30
    ).round(1)

    clientes_pendientes = ultimo_banner[
        ultimo_banner["Meses_sin_servicio"] >= meses_recordatorio
    ]

    if not clientes_pendientes.empty:
        col_banner, col_config = st.columns([4, 1])
        with col_banner:
            if st.button(
                f"🔔 {len(clientes_pendientes)} clientes llevan {meses_recordatorio}+ meses sin servicio — Ver en Follow Up →",
                use_container_width=True,
                key="btn_banner_followup"
            ):
                st.session_state["pagina"] = "Follow Up"
                st.session_state["followup_meses_override"] = meses_recordatorio
                st.rerun()
        with col_config:
            if st.button(
                "⚙️ Configurar",
                key="btn_config_recordatorio",
                use_container_width=True
            ):
                st.session_state["mostrar_config_recordatorio"] = not st.session_state.get(
                    "mostrar_config_recordatorio", False
                )

        if st.session_state.get("mostrar_config_recordatorio", False):
            nuevo_umbral = st.slider(
                "Avisar cuando un cliente lleve X meses sin servicio:",
                1, 24, meses_recordatorio,
                key="slider_recordatorio"
            )
            if nuevo_umbral != meses_recordatorio:
                st.session_state["meses_recordatorio"] = nuevo_umbral
                st.rerun()

    st.markdown("---")

    # ─────────────────────────────
    # 📊 VENTAS Y FLUJO DE EFECTIVO
    # ─────────────────────────────
    st.subheader("📊 Ventas")

    año_resumen = st.selectbox(
        "Año:",
        años_sin_2026,
        index=len(años_sin_2026)-1
    )

    df_r = df[df["Año"] == año_resumen]

    total_ventas = df_r["Monto"].sum()
    total_clientes = df_r["Nombre"].nunique()
    ticket_promedio = df_r["Monto"].mean()
    total_servicios = len(df_r)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Ventas totales", f"${total_ventas:,.0f}")
    col2.metric("Clientes únicos", f"{total_clientes:,}")
    col3.metric("Ticket promedio", f"${ticket_promedio:,.0f}")
    col4.metric("Servicios realizados", f"{total_servicios:,}")

    if año_resumen > min(años_sin_2026):
        df_ant = df[df["Año"] == año_resumen - 1]
        ventas_ant = df_ant["Monto"].sum()
        diferencia = total_ventas - ventas_ant
        porcentaje = (diferencia / ventas_ant * 100) if ventas_ant > 0 else 0
        color = "green" if diferencia > 0 else "red"
        simbolo = "▲" if diferencia > 0 else "▼"
        st.markdown(
            f"<p style='color:{color}; font-size:16px'>{simbolo} Comparado con {año_resumen-1}: "
            f"${abs(diferencia):,.0f} ({porcentaje:+.1f}%)</p>",
            unsafe_allow_html=True
        )

    # ─────────────────────────────
    # 💰 FLUJO DE EFECTIVO
    # ─────────────────────────────
    if USUARIOS[st.session_state["usuario"]].get("features", {}).get("sheets", True):
        st.subheader("💰 Flujo de efectivo")

        finanzas_usuario = USUARIOS[st.session_state["usuario"]].get("finanzas", {})

        if finanzas_usuario:
            años_finanzas = list(finanzas_usuario.keys())
            año_finanzas = st.selectbox(
                "Año financiero:",
                años_finanzas,
                index=len(años_finanzas)-1
            )
            try:
                ingresos, gastos, utilidad = cargar_finanzas(finanzas_usuario[año_finanzas])
                if ingresos is not None:
                    col1, col2, col3 = st.columns(3)
                    col1.metric("💰 Ingresos", f"${ingresos:,.0f}")
                    col2.metric("💸 Gastos", f"${gastos:,.0f}")
                    col3.metric("🟢 Utilidad", f"${utilidad:,.0f}")
                else:
                    st.warning("No se pudieron leer los datos del sheet")
            except Exception as e:
                st.error("Error cargando finanzas")
                st.write(e)
        else:
            st.info("No hay datos de flujo de efectivo configurados.")
    # ── VENTAS ──
elif pagina == "Ventas":
    st.title("Análisis de Ventas")

    _feat = USUARIOS[st.session_state["usuario"]].get("features", {})

    if _feat.get("sheets", True):
        # ── ESTADÍSTICAS FINANCIERAS DESDE SHEETS ──
        st.subheader("💰 Margen de ganancia mensual")
        finanzas_usuario = USUARIOS[st.session_state["usuario"]].get("finanzas", {})
        sheets_usuario = USUARIOS[st.session_state["usuario"]].get("sheets", {})
        nombre_hoja_stats = USUARIOS[st.session_state["usuario"]].get("estadisticas_hoja", "Estadisticas finales")
        if not finanzas_usuario:
            st.info("No hay datos financieros configurados para esta cuenta.")
        else:
            @st.cache_data(ttl=3600)
            def cargar_estadisticas_finales(_client_gs, sheet_id_stats, nombre_hoja):
                try:
                    sh = _client_gs.open_by_key(sheet_id_stats)
                    try:
                        ws = sh.worksheet(nombre_hoja)
                    except:
                        return pd.DataFrame()
                    datos = ws.get_all_values()
                    meses = []
                    for row in datos[1:13]:
                        if len(row) >= 4 and row[0].strip():
                            try:
                                ventas = float(str(row[1]).replace("$","").replace(",","").strip()) if row[1].strip() else 0
                                gastos = float(str(row[2]).replace("$","").replace(",","").strip()) if row[2].strip() else 0
                                ganancia = float(str(row[3]).replace("$","").replace(",","").strip()) if row[3].strip() and row[3].strip() != "-" else 0
                                meses.append({"Mes": row[0].strip(), "Ventas": ventas, "Gastos": gastos, "Ganancia": ganancia})
                            except:
                                pass
                    return pd.DataFrame(meses)
                except Exception as e:
                    return pd.DataFrame()

            sheet_id_stats = sheets_usuario.get(2026) or (sheets_usuario.get(max(sheets_usuario.keys())) if sheets_usuario else None)
            if sheet_id_stats:
                try:
                    client_gs_stats = get_gspread_client()
                    df_stats = cargar_estadisticas_finales(client_gs_stats, sheet_id_stats, nombre_hoja_stats)
                except:
                    df_stats = pd.DataFrame()
                if not df_stats.empty:
                    df_plot = df_stats[df_stats["Ventas"] > 0].copy()
                    if not df_plot.empty:
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Total ventas", f"${df_plot['Ventas'].sum():,.0f}")
                        col2.metric("Total gastos", f"${df_plot['Gastos'].sum():,.0f}")
                        col3.metric("Ganancia total", f"${df_plot['Ganancia'].sum():,.0f}")
                        fig_margen = px.bar(df_plot, x="Mes", y=["Ventas","Gastos","Ganancia"], barmode="group",
                            color_discrete_map={"Ventas":"#4F6AF0","Gastos":"#EF4444","Ganancia":"#10B981"},
                            labels={"value":"Pesos MXN","variable":"Concepto"})
                        fig_margen.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", legend_title_text="")
                        st.plotly_chart(fig_margen, use_container_width=True)
                        df_tabla = df_plot.copy()
                        df_tabla["Margen %"] = (df_tabla["Ganancia"] / df_tabla["Ventas"] * 100).round(1).astype(str) + "%"
                        df_tabla["Ventas"] = df_tabla["Ventas"].apply(lambda x: f"${x:,.0f}")
                        df_tabla["Gastos"] = df_tabla["Gastos"].apply(lambda x: f"${x:,.0f}")
                        df_tabla["Ganancia"] = df_tabla["Ganancia"].apply(lambda x: f"${x:,.0f}")
                        st.dataframe(df_tabla, use_container_width=True, hide_index=True)
                    else:
                        st.info("Aún no hay datos financieros para este año.")
                else:
                    st.info("No se encontró la hoja de estadísticas en el Sheet configurado.")
            else:
                st.info("No hay Sheet configurado.")
        st.markdown("---")

    elif _feat.get("gastos"):
        st.subheader("💰 Margen de ganancia mensual")
        gcli = create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])
        gcli.postgrest.auth(st.session_state.get("access_token", ""))
        try:
            gastos_resp = gcli.table("gastos").select("*").eq("empresa_id", st.session_state["empresa_id"]).execute()
            df_gastos = pd.DataFrame(gastos_resp.data) if gastos_resp.data else pd.DataFrame(columns=["id","fecha","concepto","monto","categoria"])
        except Exception as e:
            df_gastos = pd.DataFrame(columns=["id","fecha","concepto","monto","categoria"])
            st.warning(f"No se pudieron cargar los gastos: {e}")

        año_m = st.selectbox("Año:", años_sin_2026, index=len(años_sin_2026)-1, key="margen_año") if años_sin_2026 else datetime.now().year

        dfm = df[df["Año"] == año_m].copy()
        ventas_mes = dfm.groupby("Mes")["Monto"].sum()

        if not df_gastos.empty:
            df_gastos["fecha"] = pd.to_datetime(df_gastos["fecha"], errors="coerce")
            df_gastos["monto"] = pd.to_numeric(df_gastos["monto"], errors="coerce").fillna(0)
            dg = df_gastos[df_gastos["fecha"].dt.year == int(año_m)]
            gastos_mes = dg.groupby(dg["fecha"].dt.month)["monto"].sum()
        else:
            gastos_mes = pd.Series(dtype=float)

        nombres_meses = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
        filas = []
        for mnum in range(1, 13):
            v = float(ventas_mes.get(mnum, 0))
            g = float(gastos_mes.get(mnum, 0))
            if v > 0 or g > 0:
                filas.append({"Mes": nombres_meses[mnum], "Ventas": v, "Gastos": g, "Ganancia": v - g})
        df_margen = pd.DataFrame(filas)

        if not df_margen.empty:
            c1, c2, c3 = st.columns(3)
            c1.metric("Total ventas", f"${df_margen['Ventas'].sum():,.0f}")
            c2.metric("Total gastos", f"${df_margen['Gastos'].sum():,.0f}")
            c3.metric("Ganancia total", f"${df_margen['Ganancia'].sum():,.0f}")
            fig_m = px.bar(df_margen, x="Mes", y=["Ventas","Gastos","Ganancia"], barmode="group",
                color_discrete_map={"Ventas":"#4F6AF0","Gastos":"#EF4444","Ganancia":"#10B981"},
                labels={"value":"Pesos MXN","variable":"Concepto"})
            fig_m.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", legend_title_text="")
            st.plotly_chart(fig_m, use_container_width=True)
            df_mt = df_margen.copy()
            df_mt["Margen %"] = df_mt.apply(lambda r: f"{(r['Ganancia']/r['Ventas']*100):.1f}%" if r['Ventas'] > 0 else "—", axis=1)
            for c in ["Ventas","Gastos","Ganancia"]:
                df_mt[c] = df_mt[c].apply(lambda x: f"${x:,.0f}")
            st.dataframe(df_mt, use_container_width=True, hide_index=True)
        else:
            st.info("Aún no hay ventas ni gastos para este año.")

        with st.expander("➕ Registrar gasto"):
            with st.form("gasto_form"):
                gf1, gf2 = st.columns(2)
                with gf1:
                    g_fecha = st.date_input("Fecha", value=datetime.now(), key="gasto_fecha")
                    g_concepto = st.text_input("Concepto", key="gasto_concepto")
                with gf2:
                    g_monto = st.number_input("Monto", min_value=0.0, step=50.0, key="gasto_monto")
                    g_cat = st.text_input("Categoría (opcional)", key="gasto_cat")
                if st.form_submit_button("💾 Guardar gasto", use_container_width=True):
                    try:
                        gcli.table("gastos").insert({
                            "empresa_id": st.session_state["empresa_id"],
                            "fecha": g_fecha.isoformat(),
                            "concepto": g_concepto,
                            "monto": float(g_monto),
                            "categoria": g_cat or None
                        }).execute()
                        st.success("✅ Gasto registrado.")
                        st.cache_data.clear()
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")

        if not df_gastos.empty:
            with st.expander("📋 Ver / borrar gastos"):
                for _, gr in df_gastos.sort_values("fecha", ascending=False).iterrows():
                    gcol1, gcol2 = st.columns([5, 1])
                    with gcol1:
                        gfec = gr["fecha"].strftime("%d/%m/%Y") if pd.notnull(gr["fecha"]) else "—"
                        st.write(f"{gfec} — {gr.get('concepto','')} — ${float(gr.get('monto',0)):,.0f}")
                    with gcol2:
                        if st.button("🗑️", key=f"delgasto_{gr['id']}"):
                            try:
                                gcli.table("gastos").delete().eq("id", gr["id"]).execute()
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error: {e}")
        st.markdown("---")

    ventas_año = df[df["Año"].isin(años_sin_2026)].groupby("Año")["Monto"].sum().reset_index()
    ventas_año.columns = ["Año", "Total"]
    st.subheader("Ventas totales por año")
    if not ventas_año.empty:
        st.bar_chart(ventas_año.set_index("Año"))
    else:
        st.info("Aún no hay datos de ventas.")

    st.subheader("Comparación mensual")
    if años_sin_2026:
        años_sel = st.multiselect("Años:", años_sin_2026, default=años_sin_2026[-2:] if len(años_sin_2026) >= 2 else años_sin_2026)
        if años_sel:
            df_f = df[df["Año"].isin(años_sel)]
            pivot = df_f.groupby(["Año","Mes"])["Monto"].sum().reset_index()
            pivot = pivot.pivot(index="Mes", columns="Año", values="Monto").fillna(0)
            pivot = pivot.sort_index()

            nombres_meses = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                             7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
            pivot.index = pivot.index.map(nombres_meses)
            pivot_reset = pivot.reset_index()
            pivot_reset.columns.name = None

            fig = px.line(
                pivot_reset, x="Mes",
                y=[col for col in pivot_reset.columns if col != "Mes"],
                labels={"value": "Ventas", "variable": "Año"},
                category_orders={"Mes": ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]}
            )
            fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig, use_container_width=True)

            st.subheader("Proyección")
            mes_actual = datetime.now().month
            año_actual = datetime.now().year
            meses_con_datos = df[(df["Año"]==año_actual) & (df["Monto"]>0)]["Mes"].nunique()
            ventas_actual_acum = df[df["Año"]==año_actual]["Monto"].sum() if año_actual in df["Año"].values else 0
            año_anterior = año_actual - 1
            ventas_anterior_mismos = df[(df["Año"]==año_anterior) & (df["Mes"] <= mes_actual)]["Monto"].sum()
            ventas_anterior_total = df[df["Año"]==año_anterior]["Monto"].sum()

            if ventas_anterior_mismos > 0 and ventas_actual_acum > 0:
                factor = ventas_actual_acum / ventas_anterior_mismos
                proyeccion = ventas_anterior_total * factor
                tendencia = (factor - 1) * 100
                color = "green" if factor >= 1 else "red"
            else:
                proyeccion = ventas_anterior_total
                tendencia = 0
                color = "gray"

            col1, col2, col3 = st.columns(3)
            col1.metric(f"Proyección {año_actual}", f"${proyeccion:,.0f}")
            col2.metric(f"Ventas reales {año_actual}", f"${ventas_actual_acum:,.0f}")
            col3.metric(f"Tendencia vs {año_anterior}", f"{tendencia:+.1f}%")
            st.markdown(f"<p style='color:{color}'>Basado en {meses_con_datos} mes(es) de datos reales</p>", unsafe_allow_html=True)

            st.subheader(f"Detalle mes a mes — {año_actual} vs {año_anterior}")
            nombres_meses_completos = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
                                       7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
            resumen_meses = []
            for m in range(1, mes_actual + 1):
                v_ant = df[(df["Año"]==año_anterior) & (df["Mes"]==m)]["Monto"].sum()
                v_act = df[(df["Año"]==año_actual) & (df["Mes"]==m)]["Monto"].sum() if año_actual in df["Año"].values else 0
                diff = v_act - v_ant
                pct = ((diff / v_ant) * 100) if v_ant > 0 else 0
                resumen_meses.append({
                    "Mes": nombres_meses_completos[m],
                    str(año_anterior): f"${v_ant:,.0f}",
                    str(año_actual): f"${v_act:,.0f}",
                    "Diferencia": f"${diff:,.0f}",
                    "Variación": f"{pct:+.1f}%"
                })
            st.dataframe(pd.DataFrame(resumen_meses), use_container_width=True, hide_index=True)
    else:
        st.info("Aún no hay datos de ventas para mostrar.")

    st.markdown("---")
    st.subheader("📈 Conversión de Follow Up")

    if "followup_resultados" not in st.session_state:
        st.session_state["followup_resultados"] = []

    resultados = st.session_state["followup_resultados"]
    if resultados:
        import collections
        conteo_resultados = collections.Counter(r["resultado"] for r in resultados)
        total_contactados = len(resultados)
        agendaron = conteo_resultados.get("Agendó", 0)
        tasa = (agendaron / total_contactados * 100) if total_contactados > 0 else 0

        col1, col2, col3 = st.columns(3)
        col1.metric("Total contactados", total_contactados)
        col2.metric("Agendaron", agendaron)
        col3.metric("Tasa de conversión", f"{tasa:.1f}%")

        df_res = pd.DataFrame(resultados)
        conteo_df = df_res["resultado"].value_counts().reset_index()
        conteo_df.columns = ["Resultado", "Cantidad"]
        st.bar_chart(conteo_df.set_index("Resultado"))

        with st.expander("Ver detalle completo"):
            st.dataframe(
                pd.DataFrame(resultados)[["timestamp", "nombre", "resultado", "bloque"]],
                use_container_width=True, hide_index=True
            )
    else:
        st.info("Aún no hay resultados de follow up registrados.")

    _features_v = USUARIOS[st.session_state["usuario"]].get("features", {})
    if _features_v.get("vendedor") and "vendedor" in df.columns:
        st.markdown("---")
        st.subheader("👥 Ventas por vendedor")
        dfv = df.copy()
        dfv["Monto"] = pd.to_numeric(dfv["Monto"], errors="coerce").fillna(0)
        dfv["Fecha"] = pd.to_datetime(dfv["Fecha"], errors="coerce")
        colv1, colv2 = st.columns(2)
        with colv1:
            desde_v = st.date_input("Desde", value=datetime(datetime.now().year, 1, 1), key="vend_desde")
        with colv2:
            hasta_v = st.date_input("Hasta", value=datetime.now(), key="vend_hasta")
        dfv = dfv[(dfv["Fecha"].dt.date >= desde_v) & (dfv["Fecha"].dt.date <= hasta_v)]
        dfv["vendedor"] = dfv["vendedor"].fillna("Sin asignar").replace("", "Sin asignar")
        if not dfv.empty:
            resumen_v = dfv.groupby("vendedor")["Monto"].agg(["sum", "count"]).reset_index()
            resumen_v.columns = ["Vendedor", "Ventas", "Servicios"]
            resumen_v = resumen_v.sort_values("Ventas", ascending=False)
            resumen_v["Ventas"] = resumen_v["Ventas"].apply(lambda x: f"${x:,.0f}")
            st.dataframe(resumen_v, use_container_width=True, hide_index=True)
        else:
            st.info("No hay ventas en ese periodo.")

    if USUARIOS[st.session_state["usuario"]].get("features", {}).get("reporte_canales"):
        st.markdown("---")
        st.subheader("📊 Reporte de canales de venta")
        dfc = df.copy()
        dfc["Monto"] = pd.to_numeric(dfc["Monto"], errors="coerce").fillna(0)
        dfc["Fecha"] = pd.to_datetime(dfc["Fecha"], errors="coerce")
        colr1, colr2 = st.columns(2)
        with colr1:
            desde_r = st.date_input("Desde", value=datetime(datetime.now().year, 1, 1), key="rep_desde")
        with colr2:
            hasta_r = st.date_input("Hasta", value=datetime.now(), key="rep_hasta")
        dfc = dfc[(dfc["Fecha"].dt.date >= desde_r) & (dfc["Fecha"].dt.date <= hasta_r)]
        if not dfc.empty:
            dfc["canal"] = dfc["Origen"].fillna("Sin canal").replace("", "Sin canal")
            if "tipo_cliente" in dfc.columns:
                dfc["rubro"] = dfc["tipo_cliente"].fillna("Residencial").replace("", "Residencial")
            else:
                dfc["rubro"] = "Residencial"
            pivot = dfc.pivot_table(index="canal", columns="rubro", values="Monto", aggfunc="sum", fill_value=0)
            pivot["Total"] = pivot.sum(axis=1)
            pivot["Servicios"] = dfc.groupby("canal").size()
            pivot = pivot.sort_values("Total", ascending=False)
            tabla = pivot.reset_index().rename(columns={"canal": "Canal"})
            st.dataframe(tabla, use_container_width=True, hide_index=True)

            det_cols = [
                ("Fecha", "Fecha"), ("ID Cliente", "ID"), ("Nombre", "Nombre"),
                ("Tel", "Teléfono"), ("correo", "Correo"), ("Dirección", "Dirección"),
                ("Servicio", "Servicio"), ("Monto", "Monto"), ("Origen", "Canal"),
                ("tipo_cliente", "Rubro"), ("vendedor", "Vendedor"), ("estado_pago", "Estado pago"),
                ("forma_pago", "Forma pago"), ("factura", "Factura"),
                ("razon_social", "Razón social"), ("rfc", "RFC"),
            ]
            detalle = pd.DataFrame()
            for src, dst in det_cols:
                detalle[dst] = dfc[src] if src in dfc.columns else ""
            detalle["Fecha"] = pd.to_datetime(detalle["Fecha"], errors="coerce").dt.strftime("%Y-%m-%d")

            import io
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                detalle.to_excel(writer, index=False, sheet_name="Servicios")
            output.seek(0)
            st.download_button(
                "📥 Descargar Excel",
                data=output.getvalue(),
                file_name=f"canales_{desde_r}_{hasta_r}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_canales"
            )
        else:
            st.info("No hay datos en ese periodo.")
        # CLIENTES
elif pagina == "Clientes":
    import urllib.parse

    st.title("Origen de Clientes")

    año_origen = st.selectbox("Año:", años_sin_2026)
    df_o = df[df["Año"] == año_origen].copy()

    df_o["Origen"] = (
        df_o["Origen"]
        .astype(str)
        .str.strip()
        .str.lower()
    )

    origenes_config = USUARIOS[st.session_state["usuario"]].get("origenes", {})
    df_o["Origen"] = df_o["Origen"].replace(origenes_config)
    df_o["Origen"] = df_o["Origen"].replace(["", "nan"], "Sin especificar")

    origen = df_o["Origen"].value_counts().reset_index()
    origen.columns = ["Canal", "Clientes"]
    st.bar_chart(origen.set_index("Canal"))
    st.dataframe(origen, use_container_width=True)

    # 🧠 HISTORIAL
    st.markdown("### 🧠 Historial de clientes")

    df_hist = df.copy()
    df_hist["Monto"] = pd.to_numeric(df_hist["Monto"], errors="coerce")
    df_hist["Fecha"] = pd.to_datetime(df_hist["Fecha"], errors="coerce")
    df_hist["ID Cliente"] = pd.to_numeric(df_hist["ID Cliente"], errors="coerce")

    historial = df_hist.groupby("ID Cliente").agg(
        Nombre=("Nombre", "last"),
        Total_Gastado=("Monto", "sum"),
        Servicios=("Monto", "count"),
        Ultima_Visita=("Fecha", "max"),
        Ticket_Promedio=("Monto", "mean"),
        Tel=("Tel", "last"),
        Direccion=("Dirección", "last")
    ).reset_index()

    historial = historial.sort_values(by="Total_Gastado", ascending=False)

    cliente_buscar = st.text_input("🔍 Buscar cliente (nombre o ID)", key=f"buscador_historial_{año_origen}")
    if cliente_buscar:
        historial = historial[
            historial["Nombre"].str.contains(cliente_buscar, case=False, na=False) |
            historial["ID Cliente"].astype(str).str.contains(cliente_buscar, na=False)
        ]
        if cliente_buscar.isdigit():
            historial = historial.sort_values(by="ID Cliente", ascending=True)

    historial_mostrar = historial.copy()
    historial_mostrar["Total_Gastado"] = historial_mostrar["Total_Gastado"].apply(lambda x: f"${x:,.0f}")
    historial_mostrar["Ticket_Promedio"] = historial_mostrar["Ticket_Promedio"].apply(lambda x: f"${x:,.0f}" if pd.notnull(x) else "-")
    historial_mostrar["Ultima_Visita"] = historial_mostrar["Ultima_Visita"].dt.strftime("%d/%m/%Y").fillna("-")
    historial_mostrar["Direccion"] = historial_mostrar["Direccion"].fillna("-")
    historial_mostrar = historial_mostrar[[
        "ID Cliente", "Nombre", "Tel", "Direccion",
        "Total_Gastado", "Servicios", "Ultima_Visita", "Ticket_Promedio"
    ]]
    historial_mostrar.columns = [
        "ID", "Nombre", "Tel", "Dirección",
        "Total Gastado", "Servicios", "Última Visita", "Ticket Promedio"
    ]

    st.markdown("### 🏆 Top clientes")
    top_clientes = historial.head(10)
    col1, col2 = st.columns(2)
    with col1:
        st.metric("💰 Mejor cliente", top_clientes.iloc[0]["Nombre"] if not top_clientes.empty else "-")
    with col2:
        st.metric("💵 Mayor gasto", f"${top_clientes.iloc[0]['Total_Gastado']:,.0f}" if not top_clientes.empty else "$0")
    st.dataframe(historial_mostrar, use_container_width=True, hide_index=True)

    # 👤 PERFIL + EDICIÓN
    st.markdown("### 👤 Perfil del cliente")
    clientes_lista = historial["Nombre"].dropna().unique().tolist()
    cliente_sel = st.selectbox("Selecciona un cliente", clientes_lista)

    if cliente_sel:
        id_sel = historial[historial["Nombre"] == cliente_sel]["ID Cliente"].iloc[0]
        df_cliente = df_hist[df_hist["ID Cliente"] == id_sel].copy()
        df_cliente = df_cliente.sort_values(by="Fecha", ascending=False)

        total = df_cliente["Monto"].sum()
        visitas = len(df_cliente)
        ultima = df_cliente["Fecha"].max()
        promedio = df_cliente["Monto"].mean()

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total gastado", f"${total:,.0f}")
        col2.metric("Servicios", visitas)
        col3.metric("Última visita", ultima.strftime("%d/%m/%Y") if pd.notnull(ultima) else "-")
        col4.metric("Ticket promedio", f"${promedio:,.0f}")

        ultima_fila = df_cliente.iloc[0]
        tel_actual = str(ultima_fila.get("Tel", ""))
        dir_actual = str(ultima_fila.get("Dirección", ""))

        st.markdown("### ✏️ Editar datos del cliente")
        with st.form("editar_cliente_form"):
            nuevo_nombre = st.text_input("Nombre", value=cliente_sel)
            nuevo_tel = st.text_input("Teléfono", value=tel_actual if tel_actual != "nan" else "")
            nueva_dir = st.text_input("Dirección", value=dir_actual if dir_actual != "nan" else "")
            guardar = st.form_submit_button("💾 Guardar cambios", use_container_width=True)

            if guardar:
                try:
                    client_auth = create_client(
                        st.secrets["SUPABASE_URL"],
                        st.secrets["SUPABASE_KEY"]
                    )
                    client_auth.postgrest.auth(st.session_state.get("access_token", ""))
                    client_auth.table("clientes").update({
                        "nombre": nuevo_nombre,
                        "tel": nuevo_tel,
                        "direccion": nueva_dir
                    }).eq("empresa_id", st.session_state["empresa_id"])\
                      .eq("cliente_id", int(id_sel))\
                      .execute()

                    st.success(f"✅ Datos de {nuevo_nombre} actualizados en todos sus registros.")
                    st.cache_data.clear()
                    import time as t
                    t.sleep(1)
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al actualizar: {e}")

        st.markdown("### 📋 Historial completo")
        mostrar = df_cliente[[
            "Fecha", "Nombre", "Servicio", "Monto", "Origen",
            "Comentarios con llamada posterior a venta"
        ]].copy()
        mostrar.columns = ["Fecha", "Nombre", "Servicio", "Monto", "Origen", "Comentarios"]
        st.dataframe(mostrar, use_container_width=True, hide_index=True)

    # 🔴 OPORTUNIDADES DE RECUPERACIÓN
    st.markdown("## 🔴 Oportunidades de recuperación")

    df_lost = df_hist.copy()
    hoy = datetime.now()

    ultimo = df_lost.groupby("ID Cliente").agg(
        Nombre=("Nombre", "last"),
        Ultima_Visita=("Fecha", "max"),
        Total_Gastado=("Monto", "sum"),
        Tel=("Tel", "last")
    ).reset_index()
    ultimo["Meses_sin_servicio"] = ((hoy - ultimo["Ultima_Visita"]).dt.days / 30).round(1)

    col1, col2 = st.columns(2)
    with col1:
        meses_min = st.slider("Meses sin servicio", 3, 24, 6)
    with col2:
        monto_min = st.number_input("Monto mínimo ($)", value=1500)

    perdidos = ultimo[
        (ultimo["Meses_sin_servicio"] >= meses_min) &
        (ultimo["Total_Gastado"] >= monto_min)
    ]
    perdidos = perdidos.sort_values(by="Total_Gastado", ascending=False)

    col1, col2, col3 = st.columns(3)
    col1.metric("Clientes recuperables", len(perdidos))
    col2.metric("Dinero en riesgo", f"${perdidos['Total_Gastado'].sum():,.0f}")
    col3.metric(
        "Meses promedio sin servicio",
        f"{perdidos['Meses_sin_servicio'].mean():.1f}" if not perdidos.empty else "0"
    )

    st.markdown("### 📋 Lista priorizada")
    perdidos_mostrar = perdidos[["ID Cliente", "Nombre", "Tel", "Total_Gastado", "Meses_sin_servicio"]].copy()
    perdidos_mostrar["Total_Gastado"] = perdidos_mostrar["Total_Gastado"].apply(lambda x: f"${x:,.0f}")
    st.dataframe(perdidos_mostrar, use_container_width=True, hide_index=True)

    st.info("💡 Para contactar clientes ve a la página de Follow Up.")
elif pagina == "Servicios":
    st.title("Servicios más vendidos")
    import collections

    año_serv = st.selectbox("Año:", años_sin_2026)
    df_s = df[df["Año"] == año_serv].copy()
    df_s = df_s[df_s["Servicio"].notna()]
    df_s = df_s[df_s["Servicio"].astype(str).str.strip() != ""]

    categorias_config = USUARIOS[st.session_state["usuario"]].get("categorias", {})
    filas_expandidas = []
    for _, row in df_s.iterrows():
        s = str(row["Servicio"]).lower()
        encontradas = set()
        for categoria, keywords in categorias_config.items():
            for kw in keywords:
                if kw in s:
                    encontradas.add(categoria)
        if not encontradas:
            encontradas.add("Otro")
        for cat in encontradas:
            filas_expandidas.append(cat)

    conteo = collections.Counter(filas_expandidas)
    cats = pd.DataFrame(conteo.items(), columns=["Categoria", "Cantidad"])
    cats = cats.sort_values("Cantidad", ascending=False)

    fig = px.pie(cats, names="Categoria", values="Cantidad", title="Servicios más vendidos", hole=0.3)
    fig.update_traces(textposition="inside", textinfo="percent+label")
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font_color="#F0F2F6"
    )
    st.plotly_chart(fig, use_container_width=True)
    st.dataframe(cats, use_container_width=True, hide_index=True)    
    # ── FOLLOW UP ──
elif pagina == "Follow Up":
    st.title("Clientes para Follow Up")
    import urllib.parse
    import base64
    import streamlit.components.v1 as components

    if "followup_historial" not in st.session_state:
        st.session_state["followup_historial"] = []
    if "followup_resultados" not in st.session_state:
        st.session_state["followup_resultados"] = []

    df_fu = df.copy()
    df_fu["Fecha"] = pd.to_datetime(df_fu["Fecha"], errors="coerce")
    df_fu["Monto"] = pd.to_numeric(df_fu["Monto"], errors="coerce")
    df_fu["ID Cliente"] = pd.to_numeric(df_fu["ID Cliente"], errors="coerce")

    # Agrupar por ID; si no hay ID, agrupar por nombre (para no perder clientes sin ID)
    df_fu["_grp"] = df_fu.apply(
        lambda r: f"id_{int(r['ID Cliente'])}" if pd.notnull(r["ID Cliente"]) else f"nom_{str(r['Nombre']).strip().lower()}",
        axis=1
    )
    ultimo = df_fu.groupby("_grp").agg(
        IDCliente=("ID Cliente", "last"),
        Nombre=("Nombre", "last"),
        UltimoServicio=("Fecha", "max"),
        Tel=("Tel", "last"),
        Comentario=("Comentarios con llamada posterior a venta", "last")
    ).reset_index(drop=True)
    ultimo.columns = ["ID Cliente", "Nombre", "Ultimo servicio", "Tel", "Comentario"]

    meses_override = st.session_state.pop("followup_meses_override", None)

    meses_dict = {"Enero":1,"Febrero":2,"Marzo":3,"Abril":4,"Mayo":5,"Junio":6,
                  "Julio":7,"Agosto":8,"Septiembre":9,"Octubre":10,"Noviembre":11,"Diciembre":12}

    modo_fu = st.radio(
        "Filtrar clientes por:",
        ["Mes y año", "Antigüedad (tiempo sin servicio)"],
        index=1 if meses_override is not None else 0,
        horizontal=True
    )

    if modo_fu.startswith("Antigüedad"):
        # Rangos (min, max) en meses. max=None => sin límite superior.
        rangos_antiguedad = {
            "3 a 6 meses": (3, 6),
            "6 a 9 meses": (6, 9),
            "9 a 12 meses": (9, 12),
            "1 a 1.5 años": (12, 18),
            "1.5 a 2 años": (18, 24),
            "2 a 3 años": (24, 36),
            "3 a 5 años": (36, 60),
            "Más de 5 años": (60, None),
        }
        if meses_override is not None:
            min_m, max_m = meses_override, None
            st.caption(f"Sin servicio desde hace más de {min_m} meses")
            etiqueta_ant = f"+{min_m}m"
        else:
            sel_ant = st.selectbox("Sin servicio (rango de tiempo):", list(rangos_antiguedad.keys()), index=0)
            min_m, max_m = rangos_antiguedad[sel_ant]
            etiqueta_ant = sel_ant
        fecha_min = datetime.now() - timedelta(days=min_m * 30)
        if max_m is not None:
            fecha_max = datetime.now() - timedelta(days=max_m * 30)
            sin_servicio = ultimo[
                (ultimo["Ultimo servicio"] < fecha_min) &
                (ultimo["Ultimo servicio"] >= fecha_max)
            ].copy()
        else:
            sin_servicio = ultimo[ultimo["Ultimo servicio"] < fecha_min].copy()
        mes_filtro = f"Antig: {etiqueta_ant}"
    else:
        col1, col2 = st.columns(2)
        with col1:
            mes_filtro = st.selectbox("Mes del último servicio:", ["Todos","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"])
        with col2:
            años_fu = sorted(ultimo["Ultimo servicio"].dropna().dt.year.unique().astype(int).tolist(), reverse=True)
            año_filtro = st.selectbox("Año del último servicio:", ["Todos"] + [str(a) for a in años_fu])
        sin_servicio = ultimo.copy()
        if mes_filtro != "Todos":
            sin_servicio = sin_servicio[sin_servicio["Ultimo servicio"].dt.month == meses_dict[mes_filtro]]
        if año_filtro != "Todos":
            sin_servicio = sin_servicio[sin_servicio["Ultimo servicio"].dt.year == int(año_filtro)]

    # Búsqueda por nombre o ID
    buscar_followup = st.text_input("🔍 Buscar por nombre o ID", key="buscar_followup")

    if buscar_followup:
        sin_servicio = sin_servicio[
            sin_servicio["Nombre"].str.contains(buscar_followup, case=False, na=False) |
            sin_servicio["ID Cliente"].astype(str).str.contains(buscar_followup, na=False)
        ]
        if buscar_followup.isdigit():
            sin_servicio = sin_servicio.sort_values(by="ID Cliente", ascending=True)

    sin_servicio = sin_servicio.sort_values("Ultimo servicio")

    st.metric("Clientes a contactar", len(sin_servicio))
    st.dataframe(sin_servicio[["ID Cliente", "Nombre", "Tel", "Ultimo servicio", "Comentario"]], use_container_width=True, hide_index=True)

    st.markdown("### 🚀 Enviar mensaje a todos")
    plantillas_usuario = USUARIOS[st.session_state["usuario"]].get("plantillas", {})

    PLANTILLAS_MENSAJES = {
        "seguimiento": plantillas_usuario.get("seguimiento", "Hola {nombre}, te contactamos de {empresa}. Solo para dar seguimiento a tu último servicio. ¿Cómo fue tu experiencia?"),
        "recordatorio": plantillas_usuario.get("recordatorio", "Hola {nombre}, en {empresa} te recordamos que ya pasó tiempo desde tu último servicio. ¿Te gustaría agendar?"),
        "promoción": plantillas_usuario.get("promocion", "Hola {nombre}, en {empresa} tenemos una promoción especial disponible. ¿Te interesa aprovecharla?"),
        "reactivación": plantillas_usuario.get("reactivacion", "Hola {nombre}, te extrañamos en {empresa} 😄 Tenemos disponibilidad esta semana. ¿Agendamos?"),
    }
    
    plantilla_masiva = st.selectbox("Plantilla:", list(PLANTILLAS_MENSAJES.keys()), key="plantilla_masiva_followup")
    mensaje_masivo_preview = PLANTILLAS_MENSAJES[plantilla_masiva].format(nombre="[Nombre]", empresa=st.session_state.get("empresa", "nuestro negocio"))
    mensaje_masivo_edit = st.text_area("Edita el mensaje...", value=mensaje_masivo_preview, key=f"mensaje_masivo_edit_{plantilla_masiva}")

    TAMANO_BLOQUE = 20

    if not sin_servicio.empty:
        clientes_validos = sin_servicio[
            sin_servicio["Tel"].notna() &
            (sin_servicio["Tel"].astype(str).str.strip() != "") &
            (sin_servicio["Tel"].astype(str).str.strip() != "nan")
        ].reset_index(drop=True)

        total_bloques = (len(clientes_validos) + TAMANO_BLOQUE - 1) // TAMANO_BLOQUE
        st.caption(f"📦 {len(clientes_validos)} clientes divididos en {total_bloques} bloque(s) de {TAMANO_BLOQUE}")

        bloque_sel = st.selectbox(
            "Selecciona bloque a enviar:",
            [f"Bloque {i+1} ({i*TAMANO_BLOQUE+1}-{min((i+1)*TAMANO_BLOQUE, len(clientes_validos))})" for i in range(total_bloques)],
            key="bloque_sel"
        )

        idx_bloque = int(bloque_sel.split(" ")[1]) - 1
        inicio = idx_bloque * TAMANO_BLOQUE
        clientes_bloque = clientes_validos.iloc[inicio:inicio + TAMANO_BLOQUE]

        urls_bloque = []
        for _, row in clientes_bloque.iterrows():
            tel = str(row["Tel"]).replace("-", "").replace(" ", "").strip()
            tel_completo = "52" + tel
            mensaje_final = mensaje_masivo_edit.format(nombre=row["Nombre"], empresa=st.session_state.get("empresa", "nuestro negocio"))
            urls_bloque.append((row["Nombre"], f"https://wa.me/{tel_completo}?text={urllib.parse.quote(mensaje_final)}"))

        st.markdown(f"**Clientes en este bloque ({len(clientes_bloque)}):**")
        cols = st.columns(3)
        for i, (nombre_btn, url_btn) in enumerate(urls_bloque):
            with cols[i % 3]:
                st.link_button(f"💬 {nombre_btn}", url_btn)

        html_links = ""
        for nombre_link, url_link in urls_bloque:
            html_links += f'<a href="{url_link}" target="_blank" style="display:block;background-color:#25D366;color:white;text-decoration:none;padding:12px 16px;border-radius:8px;margin-bottom:8px;font-size:15px;font-family:sans-serif;">💬 {nombre_link}</a>'

        pagina_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>WhatsApp Follow Up — Bloque {idx_bloque+1}</title>
        <style>body{{font-family:sans-serif;padding:24px;max-width:500px;margin:auto;background:#f9f9f9;}}h2{{color:#128C7E;}}p{{color:#555;margin-bottom:20px;}}</style></head>
        <body><h2>📋 Bloque {idx_bloque+1} — {len(urls_bloque)} contactos</h2><p>Haz click en cada nombre para abrir WhatsApp Web en una pestaña nueva.</p>{html_links}</body></html>"""

        b64 = base64.b64encode(pagina_html.encode("utf-8")).decode("utf-8")
        components.html(f'<a href="data:text/html;base64,{b64}" target="_blank" style="display:block;background-color:#128C7E;color:white;text-decoration:none;text-align:center;border-radius:8px;padding:14px 24px;font-size:16px;font-family:sans-serif;margin-top:8px;">🚀 Abrir panel de envío — Bloque {idx_bloque+1} ({len(urls_bloque)} contactos)</a>', height=65)

        st.markdown("---")
        st.markdown("### ✅ Marcar resultado del bloque")
        st.caption("Registra qué pasó con cada cliente. Esto alimenta el dashboard de conversión en Resumen.")

        OPCIONES_RESULTADO = ["Agendó", "No contestó", "Número inválido", "No le interesa", "Pendiente"]

        with st.expander(f"Registrar resultados — Bloque {idx_bloque+1}"):
            resultados_bloque = {}
            for _, row in clientes_bloque.iterrows():
                resultados_bloque[row["Nombre"]] = st.selectbox(
                    row["Nombre"], OPCIONES_RESULTADO, index=4,
                    key=f"resultado_{row['Nombre']}_{idx_bloque}"
                )

        if st.button(f"✅ Guardar resultados y marcar bloque {idx_bloque+1} como enviado", use_container_width=True):
            timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")
            with st.spinner("Guardando..."):
                try:
                    for nombre_r, resultado_r in resultados_bloque.items():
                        st.session_state["followup_resultados"].append({
                            "nombre": nombre_r, "resultado": resultado_r,
                            "timestamp": timestamp, "bloque": idx_bloque + 1
                        })
                    st.session_state["followup_historial"].append({
                        "bloque": idx_bloque + 1, "timestamp": timestamp,
                        "mes_filtro": mes_filtro, "clientes": len(clientes_bloque),
                        "plantilla": plantilla_masiva,
                        "nombres": [n for n, _ in urls_bloque],
                        "resultados": resultados_bloque
                    })
                    st.success(f"✅ Bloque {idx_bloque+1} guardado — {timestamp}")
                    st.cache_data.clear()
                except Exception as e:
                    st.error(f"Error guardando: {e}")

    if st.session_state["followup_historial"]:
        st.markdown("### 📋 Historial de envíos esta sesión")
        for h in reversed(st.session_state["followup_historial"]):
            with st.expander(f"Bloque {h['bloque']} — {h['timestamp']} — {h['clientes']} clientes — {h['mes_filtro']}"):
                st.caption(f"Plantilla: {h['plantilla']}")
                if "resultados" in h:
                    for nombre_h, resultado_h in h["resultados"].items():
                        st.write(f"• {nombre_h}: **{resultado_h}**")
                else:
                    st.write(", ".join(h['nombres']))

    st.markdown("---")
    st.markdown("### 💬 Enviar mensaje individual")
    if not sin_servicio.empty:
        opciones_followup = sin_servicio.apply(
            lambda x: f"[{int(x['ID Cliente'])}] {x['Nombre']} - {x['Tel']}"
            if pd.notnull(x['ID Cliente']) else f"{x['Nombre']} - {x['Tel']}", axis=1
        )
        cliente_sel = st.selectbox("Selecciona cliente:", opciones_followup)
        nombre = cliente_sel.split("] ")[-1].split(" - ")[0] if "]" in cliente_sel else cliente_sel.split(" - ")[0]
        telefono = cliente_sel.split(" - ")[-1].replace("-", "").replace(" ", "")

        plantilla_ind = st.selectbox("Selecciona plantilla", list(PLANTILLAS_MENSAJES.keys()), key="plantilla_individual")
        mensaje_base = PLANTILLAS_MENSAJES[plantilla_ind]
        mensaje_generado = mensaje_base.format(nombre=nombre, empresa=st.session_state.get("empresa", "nuestro negocio"))
        mensaje = st.text_area("Mensaje", value=mensaje_generado)

        if telefono and telefono != "nan":
            telefono = "52" + telefono
            whatsapp_url = f"https://wa.me/{telefono}?text={urllib.parse.quote(mensaje)}"
            st.link_button("Enviar mensaje por WhatsApp", whatsapp_url)
        else:
            st.warning("Cliente sin teléfono válido")
# CHATBOT
elif pagina == "Chat":
    st.title("🤖 Asistente del negocio")

    st.markdown("### 💡 Ejemplos de preguntas")
    st.markdown("""
    - ¿Cuánto vendí este mes?
    - ¿Cuánto vendí el mes pasado?
    - ¿Cuánto vendí en los últimos 3 meses?
    - ¿Cuánto vendí este año?
    - ¿Quién es mi mejor cliente?
    - ¿Qué clientes no han venido en 6 meses?
    - ¿Cuántos clientes tengo?
    """)

    pregunta = st.text_input("Haz una pregunta:")

    if pregunta:

        pregunta_lower = pregunta.lower()

        hoy = datetime.now()
        mes_actual = hoy.month
        año_actual = hoy.year

        # 💰 VENTAS ESTE MES
        if "este mes" in pregunta_lower:
            df_mes = df[
                (df["Mes"] == mes_actual) &
                (df["Año"] == año_actual)
            ]
            total = df_mes["Monto"].sum()
            st.success(f"Ventas este mes ({mes_actual}/{año_actual}): ${total:,.0f}")

        # 📅 MES PASADO
        elif "mes pasado" in pregunta_lower:
            if mes_actual == 1:
                mes = 12
                año = año_actual - 1
            else:
                mes = mes_actual - 1
                año = año_actual

            df_mes = df[
                (df["Mes"] == mes) &
                (df["Año"] == año)
            ]
            total = df_mes["Monto"].sum()
            st.success(f"Ventas mes pasado ({mes}/{año}): ${total:,.0f}")

        # 📊 ÚLTIMOS 3 MESES
        elif "3 meses" in pregunta_lower:
            fecha_limite = hoy - timedelta(days=90)

            df_3m = df.copy()
            df_3m["Fecha"] = pd.to_datetime(df_3m["Fecha"], errors="coerce")

            df_3m = df_3m[df_3m["Fecha"] >= fecha_limite]

            total = df_3m["Monto"].sum()
            st.success(f"Ventas últimos 3 meses: ${total:,.0f}")

        # 📆 ESTE AÑO
        elif "este año" in pregunta_lower:
            df_año = df[df["Año"] == año_actual]
            total = df_año["Monto"].sum()
            st.success(f"Ventas {año_actual}: ${total:,.0f}")

        # 💰 VENTAS GENERALES (fallback)
        elif "ventas" in pregunta_lower:
            df_año = df[df["Año"] == año_actual]
            total = df_año["Monto"].sum()
            st.success(f"Ventas {año_actual}: ${total:,.0f}")

        # 🏆 MEJOR CLIENTE
        elif "mejor cliente" in pregunta_lower:
            if not df.empty:
                agrupado = df.groupby("Nombre")["Monto"].sum()
                top = agrupado.idxmax()
                total_top = agrupado.max()
                st.success(f"Tu mejor cliente es {top} con ${total_top:,.0f}")
            else:
                st.warning("No hay datos disponibles")

        # 👥 TOTAL CLIENTES
        elif "clientes" in pregunta_lower and "perdidos" not in pregunta_lower:
            total_clientes = df["Nombre"].nunique()
            st.success(f"Tienes {total_clientes} clientes únicos")

        # 🔴 CLIENTES PERDIDOS
        elif "perdidos" in pregunta_lower or "no han venido" in pregunta_lower:

            df_lost = df.copy()
            df_lost["Fecha"] = pd.to_datetime(df_lost["Fecha"], errors="coerce")
            df_lost["Monto"] = pd.to_numeric(df_lost["Monto"], errors="coerce")

            ultimo = df_lost.groupby("Nombre").agg(
                Ultima_Visita=("Fecha", "max"),
                Total_Gastado=("Monto", "sum"),
                Tel=("Tel", "last")
            ).reset_index()

            ultimo["Meses_sin_servicio"] = ((hoy - ultimo["Ultima_Visita"]).dt.days / 30).round(1)

            perdidos = ultimo[ultimo["Meses_sin_servicio"] >= 6]

            if not perdidos.empty:
                st.dataframe(perdidos)
            else:
                st.success("No hay clientes perdidos 🎉")

        # ❌ NO ENTENDIDO
        else:
            st.warning("Aún no entiendo esa pregunta 😅")
    # AGENDA
elif pagina == "Agenda":
    st.title("📅 Agenda de Servicios")
    import urllib.parse
    import json
    from datetime import datetime, timedelta

    SHEET_IDS = USUARIOS[st.session_state["usuario"]].get("sheets", {})
    cotizador = USUARIOS[st.session_state["usuario"]].get("cotizador", {})
    PRECIOS = cotizador.get("precios", {})
    PAQUETES_COT = cotizador.get("paquetes", [])
    MINIMO = cotizador.get("minimo", 950)
    SHEET_COLS = USUARIOS[st.session_state["usuario"]].get("sheet_columnas", {})
    features = USUARIOS[st.session_state["usuario"]].get("features", {})
    def get_supabase_auth():
        client_auth = create_client(
            st.secrets["SUPABASE_URL"],
            st.secrets["SUPABASE_KEY"]
        )
        client_auth.postgrest.auth(st.session_state.get("access_token", ""))
        return client_auth

    def limpiar_valor(val, default=""):
        if val is None:
            return default
        try:
            if pd.isna(val):
                return default
        except:
            pass
        return val

    def fecha_relativa(fecha):
        hoy = datetime.now().date()
        diff = (fecha - hoy).days
        dias_es = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
        if diff == 0:
            return "hoy"
        elif diff == 1:
            return "mañana"
        elif diff == 2:
            return "pasado mañana"
        elif 3 <= diff <= 6:
            return f"el {dias_es[fecha.weekday()]}"
        else:
            return fecha.strftime("%d/%m/%Y")

    def rango_hora(hora_str):
        if not hora_str:
            return ""
        try:
            h = datetime.strptime(str(hora_str).strip(), "%H:%M")
            h2 = h + timedelta(minutes=30)
            return f"{h.strftime('%H:%M')} - {h2.strftime('%H:%M')}"
        except:
            return str(hora_str)

    def get_precio(servicio, paquete, cantidad=1):
        try:
            p = PRECIOS.get(servicio, {}).get(paquete, 0)
            return p * cantidad
        except:
            return 0

    def construir_fila_sheet(siguiente_folio, folio_interno, fecha_str, id_cli_limpio,
                              nombre, tel, direccion, origen, monto, servicio, comentarios):
        # Número de columnas total
        max_col = max(SHEET_COLS.values()) + 1 if SHEET_COLS else 14
        fila = [""] * max(max_col, 14)

        fila[SHEET_COLS.get("folio_sistema", 0)] = siguiente_folio
        fila[SHEET_COLS.get("folio_interno", 1)] = folio_interno
        fila[SHEET_COLS.get("fecha", 2)] = fecha_str
        fila[SHEET_COLS.get("id_cliente", 3)] = id_cli_limpio
        fila[SHEET_COLS.get("nombre", 4)] = nombre
        fila[SHEET_COLS.get("tel", 5)] = tel
        fila[SHEET_COLS.get("direccion", 6)] = direccion
        fila[SHEET_COLS.get("origen", 7)] = origen
        fila[SHEET_COLS.get("monto", 8)] = monto
        fila[SHEET_COLS.get("servicio", 9)] = servicio
        fila[SHEET_COLS.get("comentarios", 10)] = comentarios

        return fila

    df_a = df.copy()
    df_a["Fecha"] = pd.to_datetime(df_a["Fecha"], errors="coerce")
    df_a["Monto"] = pd.to_numeric(df_a["Monto"], errors="coerce")
    if "ID Cliente" not in df_a.columns:
        df_a["ID Cliente"] = None
    if "hora" not in df_a.columns:
        df_a["hora"] = None
    df_a["ID Cliente"] = pd.to_numeric(df_a["ID Cliente"], errors="coerce")

    plantillas = USUARIOS[st.session_state["usuario"]].get("plantillas", {})
    empresa = st.session_state.get("empresa", "")

    # 💰 Banner de cuentas por cobrar
    if features.get("cobranza") and not df_a.empty and "estado_pago" in df_a.columns:
        _pc = df_a[(df_a["realizado"] == True) & (df_a["estado_pago"].isin(["pendiente", "parcial"]))].copy()
        if not _pc.empty:
            _pc["Monto"] = pd.to_numeric(_pc["Monto"], errors="coerce").fillna(0)
            _pc["_pagado"] = pd.to_numeric(_pc["monto_pagado"], errors="coerce").fillna(0) if "monto_pagado" in _pc.columns else 0
            _pc["_saldo"] = _pc["Monto"] - _pc["_pagado"]
            total_saldo = _pc["_saldo"].sum()
            st.warning(f"💰 **{len(_pc)} cuenta(s) por cobrar** — saldo total: ${total_saldo:,.0f}")
            with st.expander("Ver cuentas por cobrar"):
                st.caption("Elige la forma de pago y marca como pagado. Para pagos parciales, abre el servicio en el calendario.")
                for _, r in _pc.sort_values("Fecha").iterrows():
                    f = r["Fecha"].date().strftime("%d/%m/%Y") if pd.notnull(r["Fecha"]) else "—"
                    est = r.get("estado_pago") or "pendiente"
                    rid = str(r.get("id", ""))
                    cbc1, cbc2, cbc3 = st.columns([3, 1.4, 1])
                    with cbc1:
                        st.write(f"**{limpiar_valor(r.get('Nombre'))}** — {f} — saldo ${r['_saldo']:,.0f} ({est})")
                    with cbc2:
                        forma_sel = st.selectbox(
                            "Forma", ["Efectivo", "Transferencia", "Tarjeta", "Cheque"],
                            key=f"forma_{rid}", label_visibility="collapsed"
                        )
                    with cbc3:
                        if rid and st.button("✅ Pagado", key=f"cobrar_{rid}"):
                            try:
                                client_auth = get_supabase_auth()
                                client_auth.table("clientes").update({
                                    "estado_pago": "pagado",
                                    "forma_pago": forma_sel,
                                    "fecha_pago": datetime.now().date().isoformat(),
                                    "monto_pagado": float(r["Monto"])
                                }).eq("id", rid).execute()
                                st.success(f"✅ {limpiar_valor(r.get('Nombre'))} pagado ({forma_sel}).")
                                st.cache_data.clear()
                                import time as t
                                t.sleep(1)
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error: {e}")
    fecha_sel = st.date_input("Selecciona una fecha", datetime.now(), key="agenda_fecha_1")
    df_dia = df_a[df_a["Fecha"].dt.date == fecha_sel]

    total_dia = df_dia["Monto"].sum()
    col1, col2 = st.columns(2)
    col1.metric("Servicios ese día", len(df_dia))
    col2.metric("Ingresos del día", f"${total_dia:,.0f}")
    st.markdown("---")

    if df_dia.empty:
        st.info("No hay servicios agendados para este día.")
    else:
        for _, row in df_dia.iterrows():
            with st.container():
                realizado = row.get("realizado", False)
                estado = "✅ Realizado" if realizado else "⏳ Pendiente"
                hora_str = limpiar_valor(row.get("hora"))
                fecha_row = row["Fecha"].date() if pd.notnull(row["Fecha"]) else None
                fecha_txt = fecha_relativa(fecha_row) if fecha_row else ""
                rango = rango_hora(hora_str)

                st.markdown(f"""
                **👤 Cliente:** {limpiar_valor(row.get('Nombre'))}  
                **🆔 ID:** {limpiar_valor(row.get('ID Cliente'))}  
                **📞 Tel:** {limpiar_valor(row.get('Tel'))}  
                **📍 Dirección:** {limpiar_valor(row.get('Dirección'))}  
                **🧼 Servicio:** {limpiar_valor(row.get('Servicio'))}  
                **📅 Fecha:** {fecha_row.strftime('%d/%m/%Y') if fecha_row else ''}{f' entre las {rango}' if rango else ''}  
                **💰 Monto:** ${row.get('Monto', 0) or 0:,.0f}  
                **Estado:** {estado}
                """)
                tel = str(row.get("Tel", "")).replace("-", "").replace(" ", "")
                if tel and tel != "nan":
                    tel_completo = "52" + tel
                    hora_msg = f" entre las {rango}" if rango else ""
                    fecha_completa = f"{fecha_row.strftime('%d/%m/%Y')} ({fecha_txt})" if fecha_row else ""
                    mensaje_confirmacion = plantillas.get("confirmacion", "Hola {nombre}, confirmamos tu servicio con {empresa} para el {fecha}{hora}.").format(
                        nombre=limpiar_valor(row.get("Nombre")),
                        empresa=empresa,
                        fecha=fecha_completa,
                        hora=hora_msg
                    )
                    url = f"https://wa.me/{tel_completo}?text={urllib.parse.quote(mensaje_confirmacion)}"
                    st.markdown(f"[💬 Enviar WhatsApp]({url})")
                else:
                    st.warning("Cliente sin teléfono")
                st.markdown("---")

    st.markdown("### ➕ Agendar nuevo servicio")

    if "form_key" not in st.session_state:
        st.session_state["form_key"] = 0

    if st.button("🗑️ Vaciar campos", key="vaciar_campos"):
        st.session_state["form_key"] += 1
        st.session_state["agenda_limpiar"] = True
        for k in ["agenda_cliente_id", "agenda_cliente_nombre", "agenda_cliente_tel", "agenda_cliente_dir"]:
            st.session_state.pop(k, None)
        st.rerun()

    limpiar = st.session_state.pop("agenda_limpiar", False)

    clientes_info = df_a.groupby("ID Cliente").agg(
        Nombre=("Nombre", "last"),
        Telefonos=("Tel", lambda x: " / ".join(
            str(t) for t in x.dropna().unique()
            if str(t).strip() not in ["", "nan"]
        )),
        Direccion=("Dirección", "last"),
    ).reset_index() if not df_a.empty and df_a["ID Cliente"].notna().any() else pd.DataFrame(columns=["ID Cliente", "Nombre", "Telefonos", "Direccion"])

    clientes_info["ID Cliente"] = pd.to_numeric(clientes_info["ID Cliente"], errors="coerce")
    clientes_info = clientes_info.dropna(subset=["ID Cliente"])
    clientes_info["ID Cliente"] = clientes_info["ID Cliente"].astype(int)

    buscar_agenda = st.text_input(
        "🔍 Escribe nombre o ID para buscar cliente",
        key=f"buscar_agenda_{st.session_state['form_key']}"
    )

    if buscar_agenda and not limpiar and not clientes_info.empty:
        if buscar_agenda.strip().isdigit():
            clientes_filtrados = clientes_info[
                clientes_info["ID Cliente"].astype(str).str.startswith(buscar_agenda.strip())
            ].sort_values("ID Cliente", ascending=True)
        else:
            clientes_filtrados = clientes_info[
                clientes_info["Nombre"].str.contains(buscar_agenda.strip(), case=False, na=False)
            ].sort_values("Nombre")

        st.caption(f"{len(clientes_filtrados)} resultado(s)")

        for _, fila in clientes_filtrados.head(10).iterrows():
            label = f"[{int(fila['ID Cliente'])}] {fila['Nombre']} — {fila['Telefonos']}"
            if st.button(label, key=f"sel_cliente_{int(fila['ID Cliente'])}_{st.session_state['form_key']}"):
                fk = st.session_state['form_key']
                dir_val = str(fila["Direccion"]) if pd.notnull(fila["Direccion"]) else ""
                st.session_state["agenda_cliente_id"] = int(fila["ID Cliente"])
                st.session_state["agenda_cliente_nombre"] = fila["Nombre"]
                st.session_state["agenda_cliente_tel"] = fila["Telefonos"]
                st.session_state["agenda_cliente_dir"] = dir_val
                # Escribir directo en las keys de los widgets para que SÍ se muestren
                st.session_state[f"nombre_{fk}"] = fila["Nombre"]
                st.session_state[f"tel_{fk}"] = fila["Telefonos"]
                st.session_state[f"dir_{fk}"] = dir_val
                st.rerun()

    id_cliente_default = None
    nombre_default = ""
    tel_default = ""
    dir_default = ""

    if "agenda_cliente_id" in st.session_state and not limpiar:
        id_cliente_default = st.session_state["agenda_cliente_id"]
        nombre_default = st.session_state.get("agenda_cliente_nombre", "")
        tel_default = st.session_state.get("agenda_cliente_tel", "")
        dir_default = st.session_state.get("agenda_cliente_dir", "")
        st.info(f"✅ Cliente seleccionado: [{id_cliente_default}] {nombre_default}")

    ORIGENES = list(USUARIOS[st.session_state["usuario"]].get("origenes", {}).keys()) or ["Rep", "Int", "Rec"]
    PAQUETES = [""] + PAQUETES_COT
    SERVICIOS_LISTA = [""] + list(PRECIOS.keys()) + ["Otro"]
    N_RENGLONES = 8

    nombre = st.text_input("Nombre del cliente", key=f"nombre_{st.session_state['form_key']}")
    telefono = st.text_input("Teléfono(s)", key=f"tel_{st.session_state['form_key']}")
    direccion = st.text_input("Dirección", key=f"dir_{st.session_state['form_key']}")
    correo_in = st.text_input("Correo (opcional)", key=f"correo_{st.session_state['form_key']}") 
    tipo_cliente_sel = "Residencial"
    razon_social_in = ""
    rfc_in = ""
    if features.get("comercial"):
        tipo_cliente_sel = st.selectbox("Tipo de cliente", ["Residencial", "Comercial"], key=f"tipocli_{st.session_state['form_key']}")
        if tipo_cliente_sel == "Comercial":
            colf1, colf2 = st.columns(2)
            with colf1:
                razon_social_in = st.text_input("Razón social", key=f"razon_{st.session_state['form_key']}")
            with colf2:
                rfc_in = st.text_input("RFC", key=f"rfc_{st.session_state['form_key']}")
    st.markdown("**🧼 Servicios**")
    st.caption("El precio se autocompleta desde el cotizador, pero puedes editarlo")

    h1, h2, h3, h4, h5 = st.columns([4, 1, 2, 2, 2])
    h1.markdown("**Servicio**")
    h2.markdown("**Cant.**")
    h3.markdown("**Paquete**")
    h4.markdown("**P. Unit.**")
    h5.markdown("**Subtotal**")

    renglones = []
    for i in range(N_RENGLONES):
        c1, c2, c3, c4, c5 = st.columns([4, 1, 2, 2, 2])
        with c1:
            serv_i = st.selectbox(f"s{i}", SERVICIOS_LISTA, label_visibility="collapsed", key=f"serv_{i}_{st.session_state['form_key']}")
            desc_otro = ""
            if serv_i == "Otro":
                desc_otro = st.text_input(
                    f"desc{i}", placeholder="Ej: carriola, cojín, base de cama, cama de perro...",
                    label_visibility="collapsed", key=f"descotro_{i}_{st.session_state['form_key']}"
                )
        with c2:
            cant_i = st.number_input(
                f"c{i}", min_value=0.0, value=0.0, step=0.5, format="%.2f",
                label_visibility="collapsed", key=f"cant_{i}_{st.session_state['form_key']}"
            )
        with c3:
            paq_i = st.selectbox(f"p{i}", PAQUETES, label_visibility="collapsed", key=f"paq_{i}_{st.session_state['form_key']}")

        # Precio sugerido desde el cotizador
        precios_serv = PRECIOS.get(serv_i, {}) if (serv_i and serv_i != "Otro") else {}
        if paq_i and paq_i in precios_serv:
            precio_default = precios_serv[paq_i]
        elif precios_serv and len(set(precios_serv.values())) == 1:
            # Servicio de precio único (PURT, Aspirado, Piel): mismo precio sin importar paquete
            precio_default = next(iter(precios_serv.values()))
        else:
            precio_default = 0.0

        with c4:
            p_unit = st.number_input(
                f"pu{i}", min_value=0.0, value=float(precio_default), step=1.0,
                label_visibility="collapsed",
                key=f"punit_{i}_{serv_i}_{paq_i}_{st.session_state['form_key']}"
            )
        subtotal_i = p_unit * cant_i if (serv_i and cant_i > 0) else 0
        with c5:
            st.markdown(f"**${subtotal_i:,.0f}**" if subtotal_i else "—")

        if serv_i:
            cant_val = int(cant_i) if float(cant_i).is_integer() else round(cant_i, 2)
            nombre_servicio = desc_otro.strip() if (serv_i == "Otro" and desc_otro.strip()) else serv_i
            renglones.append({
                "servicio": nombre_servicio,
                "cantidad": cant_val,
                "paquete": paq_i,
                "precio_unitario": p_unit,
                "subtotal": subtotal_i
            })

    st.markdown("---")
    hay_por_cotizar = any(r["precio_unitario"] is None for r in renglones)
    subtotal_total = None if hay_por_cotizar else sum((r["subtotal"] or 0) for r in renglones)

    col_opciones, col_totales = st.columns([2, 1])
    with col_opciones:
        aplica_descuento = st.checkbox("Aplicar descuento", key=f"desc_{st.session_state['form_key']}")
        descuento_pct = st.number_input("Descuento (%)", min_value=0, max_value=100, value=0, key=f"desc_pct_{st.session_state['form_key']}") if aplica_descuento else 0
        aplica_iva = st.checkbox("Aplicar IVA 16% (factura)", key=f"iva_{st.session_state['form_key']}")
        sin_cotizar = st.checkbox("🖨️ Imprimir sin cotizar (total en blanco)", key=f"sincot_{st.session_state['form_key']}")
        solicitar_anticipo = False
        monto_anticipo_in = 0.0
        if features.get("anticipo"):
            solicitar_anticipo = st.checkbox("🔖 Solicitar anticipo (cliente nuevo)", key=f"anticipo_{st.session_state['form_key']}")
            monto_anticipo_in = st.number_input("Monto de anticipo", min_value=0.0, step=50.0, key=f"monto_anticipo_{st.session_state['form_key']}") if solicitar_anticipo else 0.0
        minimo_manual = st.number_input("Mínimo manual (opcional)", min_value=0.0, value=0.0, step=50.0, key=f"minman_{st.session_state['form_key']}") if features.get("minimo_manual") else 0.0

    # Mínimo efectivo = el mayor entre el de config y el manual
    min_efectivo = max(MINIMO, minimo_manual)
    if hay_por_cotizar:
        subtotal_final = None
    else:
        subtotal_final = max(subtotal_total, min_efectivo) if subtotal_total > 0 else 0

    if subtotal_final is None:
        monto_descuento = 0
        iva = 0
        total_final = None
    else:
        monto_descuento = subtotal_final * descuento_pct / 100
        base = subtotal_final - monto_descuento
        iva = base * 0.16 if aplica_iva else 0
        total_final = base + iva

    with col_totales:
        if subtotal_final is None:
            st.warning("⚠️ Cotización incompleta: hay servicios *por cotizar*. No se calcula el total hasta que tengan precio.")
        elif subtotal_final > 0:
            st.markdown(f"**Subtotal:** ${subtotal_final:,.0f}")
            if subtotal_total is not None and 0 < subtotal_total < min_efectivo:
                st.caption(f"⚠️ Aplicando mínimo ${min_efectivo:,.0f}")
            if aplica_descuento and descuento_pct > 0:
                st.markdown(f"**Descuento ({descuento_pct}%):** -${monto_descuento:,.0f}")
            if aplica_iva:
                st.markdown(f"**IVA 16%:** ${iva:,.0f}")
            st.markdown(f"### 💰 Total: ${total_final:,.0f}")
            if sin_cotizar:
                st.warning("⚠️ Hoja SIN COTIZAR: el total saldrá en blanco en el PDF. Cotiza en sitio y anota el total a mano.")
    origen_input = st.selectbox("Origen", ORIGENES, key=f"origen_{st.session_state['form_key']}")
    vendedor_in = st.text_input("Vendedor (opcional)", key=f"vendedor_{st.session_state['form_key']}", placeholder="Ej. Hermana") if features.get("vendedor") else ""
    col1, col2 = st.columns(2)
    with col1:
        fecha = st.date_input("Fecha", datetime.now(), key=f"fecha_{st.session_state['form_key']}")
    with col2:
        hora = st.time_input("Hora", value=None, key=f"hora_{st.session_state['form_key']}")

    if st.button("✅ Agendar", use_container_width=True, type="primary", key=f"btn_agendar_{st.session_state['form_key']}"):
        if not nombre.strip():
            st.error("El nombre del cliente es obligatorio.")
        elif not any(r["servicio"] for r in renglones):
            st.error("Agrega al menos un servicio.")
        else:
            conflicto = False
            if hora:
                hora_str_nueva = hora.strftime("%H:%M")
                df_conflicto = df_a[
                    (df_a["Fecha"].dt.date == fecha) &
                    (df_a["hora"] == hora_str_nueva) &
                    (df_a["realizado"] == False)
                ]
                if not df_conflicto.empty:
                    conflicto = True
                    nombres_conflicto = ", ".join(df_conflicto["Nombre"].dropna().tolist())
                    st.warning(f"⚠️ Ya hay un servicio a las {hora_str_nueva} el {fecha.strftime('%d/%m/%Y')} — {nombres_conflicto}.")

            if not conflicto or st.session_state.get("forzar_agenda", False):
                try:
                    client_auth = get_supabase_auth()

                    id_cliente = None
                    if id_cliente_default and str(id_cliente_default) not in ["", "nan", "None"]:
                        id_cliente = int(id_cliente_default)
                    else:
                        # Evitar duplicados: si ya existe un cliente con ese teléfono, reusar su ID
                        if telefono and str(telefono).strip().lower() not in ["", "nan", "none"]:
                            existe = client_auth.table("clientes")\
                                .select("cliente_id")\
                                .eq("empresa_id", st.session_state["empresa_id"])\
                                .eq("tel", telefono)\
                                .not_.is_("cliente_id", "null")\
                                .order("cliente_id", desc=True)\
                                .limit(1)\
                                .execute()
                            if existe.data and existe.data[0]["cliente_id"]:
                                id_cliente = int(existe.data[0]["cliente_id"])
                        if id_cliente is None:
                            max_id_resp = client_auth.table("clientes")\
                                .select("cliente_id")\
                                .eq("empresa_id", st.session_state["empresa_id"])\
                                .not_.is_("cliente_id", "null")\
                                .order("cliente_id", desc=True)\
                                .limit(1)\
                                .execute()
                            if max_id_resp.data and max_id_resp.data[0]["cliente_id"]:
                                id_cliente = int(max_id_resp.data[0]["cliente_id"]) + 1
                            else:
                                id_cliente = 1

                    hora_str = hora.strftime("%H:%M") if hora else None
                    rango_nuevo = rango_hora(hora_str)
                    renglones_validos = [r for r in renglones if r["servicio"]]
                    servicio_str = " | ".join([f"{r['cantidad']} {r['servicio']} {r['paquete']}" for r in renglones_validos])
                    renglones_json = json.dumps(renglones_validos, ensure_ascii=False)

                    client_auth.table("clientes").insert({
                        "empresa_id": st.session_state["empresa_id"],
                        "cliente_id": id_cliente,
                        "nombre": nombre,
                        "tel": telefono,
                        "direccion": direccion,
                        "servicio": servicio_str,
                        "cantidad": str(len(renglones_validos)),
                        "paquete": renglones_validos[0]["paquete"] if renglones_validos else "",
                        "fecha": fecha.isoformat(),
                        "monto": float(total_final) if total_final is not None else 0,
                        "origen": origen_input,
                        "año": fecha.year,
                        "realizado": False,
                        "hora": hora_str,
                        "comentarios": renglones_json,
                        "descuento_pct": float(descuento_pct),
                        "iva": bool(aplica_iva),
                        "estado_pago": "pendiente",
                        "tipo_cliente": tipo_cliente_sel,
                        "razon_social": razon_social_in or None,
                        "rfc": rfc_in or None,
                        "anticipo_solicitado": bool(solicitar_anticipo),
                        "monto_anticipo": float(monto_anticipo_in),
                        "vendedor": vendedor_in or None,
                        "correo": correo_in or None
                    }).execute()
                    fecha_txt = fecha_relativa(fecha)
                    hora_txt = f" entre las {rango_nuevo}" if rango_nuevo else ""
                    st.success(f"✅ Agendado para {nombre} (ID: {id_cliente}) — {fecha.strftime('%d/%m/%Y')} ({fecha_txt}){hora_txt} — Total: ${total_final:,.0f}")

                    try:
                        if sin_cotizar or total_final is None:
                            h_sub, h_desc, h_iva, h_tot, h_dpct = 0, 0, 0, 0, 0
                        else:
                            h_sub, h_desc, h_iva, h_tot, h_dpct = subtotal_final, monto_descuento, iva, total_final, descuento_pct
                        pdf_bytes = generar_hoja_servicio(
                            nombre=nombre,
                            direccion=direccion,
                            telefono=telefono,
                            fecha=fecha.strftime("%d/%m/%Y"),
                            hora=rango_nuevo,
                            folio="",
                            origen=origen_input,
                            items=renglones_validos,
                            subtotal=h_sub,
                            descuento=h_desc,
                            descuento_pct=h_dpct,
                            iva=h_iva,
                            total=h_tot,
                            ciudad=USUARIOS[st.session_state["usuario"]].get("ciudad", ""),
                            template_path=USUARIOS[st.session_state["usuario"]].get("template_pdf") or "assets/Hoja de servicio de Maxi Clean.pdf"
                        )
                        st.download_button(
                            "📄 Descargar hoja de servicio",
                            data=pdf_bytes,
                            file_name=f"hoja_{nombre.replace(' ','_')}_{fecha.strftime('%d%m%Y')}.pdf",
                            mime="application/pdf",
                            key="download_hoja_nueva"
                        )
                    except Exception as e_pdf:
                        st.warning(f"No se pudo generar la hoja: {e_pdf}")

                    st.cache_data.clear()
                    st.session_state.pop("forzar_agenda", None)
                    for k in ["agenda_cliente_id", "agenda_cliente_nombre", "agenda_cliente_tel", "agenda_cliente_dir"]:
                        st.session_state.pop(k, None)
                    if "evento_seleccionado" in st.session_state:
                        del st.session_state["evento_seleccionado"]
                    import time as t
                    t.sleep(1)
                    st.rerun()

                except Exception as e:
                    st.error(f"Error al agendar: {e}")
            else:
                st.session_state["forzar_agenda"] = True

    if st.session_state.get("forzar_agenda", False):
        if st.button("⚠️ Agendar de todas formas", use_container_width=True):
            st.rerun()

    st.markdown("---")

    from streamlit_calendar import calendar
    st.markdown("### 📅 Calendario de servicios")

    hoy = datetime.now()
    inicio_cal = hoy.replace(day=1) - pd.DateOffset(months=1)
    fin_cal = hoy + pd.DateOffset(months=12)
    df_cal = df_a[
        (df_a["Fecha"] >= inicio_cal) &
        (df_a["Fecha"] <= fin_cal)
    ].copy() if not df_a.empty else pd.DataFrame()
    if not df_cal.empty:
        df_cal = df_cal.dropna(subset=["Fecha"])

    eventos = []
    for _, row in df_cal.iterrows():
        realizado = row.get("realizado", False)
        color = "#4F6AF0" if realizado else "#F59E0B"
        hora_ev = limpiar_valor(row.get("hora"))
        rango_ev = rango_hora(hora_ev)
        titulo = f"{'✅' if realizado else '⏳'} {limpiar_valor(row.get('Nombre'))} — {limpiar_valor(row.get('Servicio'))}"
        if rango_ev:
            titulo += f" {rango_ev}"
        eventos.append({
            "title": titulo,
            "start": row["Fecha"].strftime("%Y-%m-%d"),
            "end": row["Fecha"].strftime("%Y-%m-%d"),
            "backgroundColor": color,
            "borderColor": color,
            "extendedProps": {"id": str(row.get("id", ""))},
        })
    opciones_calendario = {
        "initialView": "dayGridMonth",
        "locale": "es",
        "headerToolbar": {
            "left": "prev,next today",
            "center": "title",
            "right": "dayGridMonth,timeGridWeek,listWeek"
        },
        "height": 600,
    }

    import hashlib
    cal_sig = hashlib.md5(str(eventos).encode()).hexdigest()[:8]
    resultado_cal = calendar(
        events=eventos,
        options=opciones_calendario,
        key=f"calendario_principal_{cal_sig}"
    )

    if resultado_cal and resultado_cal.get("eventClick"):
        evento_click = resultado_cal["eventClick"]["event"]
        id_click = evento_click.get("extendedProps", {}).get("id")
        if id_click:
            st.session_state["evento_seleccionado"] = {"id": id_click}

    if "evento_seleccionado" in st.session_state:
        ev = st.session_state["evento_seleccionado"]
        id_click = ev["id"]
        df_evento = df_cal[df_cal["id"].astype(str) == str(id_click)] if not df_cal.empty else pd.DataFrame()

        if not df_evento.empty:
            row = df_evento.iloc[0]
            realizado = row.get("realizado", False)
            hora_row = limpiar_valor(row.get("hora"))
            rango_row = rango_hora(hora_row)
            fecha_row = row["Fecha"].date() if pd.notnull(row["Fecha"]) else None
            fecha_txt = fecha_relativa(fecha_row) if fecha_row else ""

            st.markdown("---")
            st.markdown("### 📋 Detalle del servicio")

            col1, col2 = st.columns([3, 1])
            with col1:
                st.markdown(f"**👤 Cliente:** {limpiar_valor(row.get('Nombre'))}")
                st.markdown(f"**🆔 ID:** {limpiar_valor(row.get('ID Cliente'))}")
                st.markdown(f"**📞 Tel:** {limpiar_valor(row.get('Tel'))}")
                st.markdown(f"**📍 Dirección:** {limpiar_valor(row.get('Dirección'))}")
                st.markdown(f"**🧼 Servicio:** {limpiar_valor(row.get('Servicio'))}")
                st.markdown(f"**📅 Fecha:** {fecha_row.strftime('%d/%m/%Y') if fecha_row else ''}{f' entre las {rango_row}' if rango_row else ''}")
                st.markdown(f"**💰 Monto:** ${row.get('Monto', 0) or 0:,.0f}")
                st.markdown(f"**🔍 Origen:** {limpiar_valor(row.get('Origen'))}")
                if features.get("cobranza"):
                    st.markdown(f"**💳 Pago:** {limpiar_valor(row.get('estado_pago'), 'sin registrar')}")
                if features.get("comercial"):
                    _linea_tipo = f"**🏢 Tipo:** {limpiar_valor(row.get('tipo_cliente'), 'Residencial')}"
                    if limpiar_valor(row.get('factura')):
                        _linea_tipo += f" · 🧾 {limpiar_valor(row.get('factura'))}"
                    st.markdown(_linea_tipo)
                if features.get("anticipo") and row.get('anticipo_solicitado'):
                    st.markdown(f"**🔖 Anticipo:** ${float(row.get('monto_anticipo') or 0):,.0f} solicitado")
                st.markdown(f"**Estado:** {'✅ Realizado' if realizado else '⏳ Pendiente'}")
            with col2:
                if st.button("✖ Cerrar detalle", use_container_width=True):
                    del st.session_state["evento_seleccionado"]
                    st.rerun()

            tel_ev = str(row.get("Tel", "")).replace("-", "").replace(" ", "")
            if tel_ev and tel_ev not in ["nan", ""]:
                tel_ev_completo = "52" + tel_ev
                if not realizado:
                    hora_msg = f" entre las {rango_row}" if rango_row else ""
                    fecha_completa_ev = f"{fecha_row.strftime('%d/%m/%Y')} ({fecha_txt})" if fecha_row else ""
                    mensaje_confirmacion = plantillas.get("confirmacion", "Hola {nombre}, confirmamos tu servicio con {empresa} para el {fecha}{hora}.").format(
                        nombre=limpiar_valor(row.get("Nombre")),
                        empresa=empresa,
                        fecha=fecha_completa_ev,
                        hora=hora_msg
                    )
                    url_confirmacion = f"https://wa.me/{tel_ev_completo}?text={urllib.parse.quote(mensaje_confirmacion)}"
                    st.markdown(f"[💬 Enviar recordatorio]({url_confirmacion})")
                else:
                    msg_sat = plantillas.get("satisfaccion", "Hola {nombre}, gracias por confiar en {empresa} 😊 ¿Cómo quedó tu servicio?").format(
                        nombre=limpiar_valor(row.get("Nombre")),
                        empresa=empresa
                    )
                    url_sat = f"https://wa.me/{tel_ev_completo}?text={urllib.parse.quote(msg_sat)}"
                    st.link_button("💬 Enviar WhatsApp de satisfacción", url_sat)

            try:
                folio_ev = limpiar_valor(row.get("folio")) or ""
                comentarios_raw = limpiar_valor(row.get("Comentarios con llamada posterior a venta"))
                try:
                    items_ev = json.loads(comentarios_raw) if comentarios_raw and comentarios_raw.startswith("[") else None
                except:
                    items_ev = None

                if items_ev:
                    hay_por_cotizar_ev = any(r.get("precio_unitario") is None or r.get("subtotal") is None for r in items_ev)
                    subtotal_ev = None if hay_por_cotizar_ev else max(sum((r.get("subtotal") or 0) for r in items_ev), MINIMO)
                else:
                    items_ev = [{"servicio": limpiar_valor(row.get("Servicio")), "cantidad": 1, "paquete": "", "precio_unitario": None, "subtotal": row.get("Monto", 0) or 0}]
                    subtotal_ev = row.get("Monto", 0) or 0

                # Descuento e IVA guardados al agendar
                try:
                    desc_pct_ev = float(row.get("descuento_pct", 0) or row.get("Descuento_pct", 0) or 0)
                except:
                    desc_pct_ev = 0
                iva_raw = row.get("iva", row.get("IVA", False))
                aplica_iva_ev = str(iva_raw).strip().lower() in ["true", "1", "t", "yes", "si", "sí"]

                if subtotal_ev is None:
                    monto_desc_ev, iva_ev, total_ev, desc_pct_ev = 0, 0, None, 0
                else:
                    monto_desc_ev = subtotal_ev * desc_pct_ev / 100
                    base_ev = subtotal_ev - monto_desc_ev
                    iva_ev = base_ev * 0.16 if aplica_iva_ev else 0
                    total_ev = base_ev + iva_ev

                pdf_bytes_ev = generar_hoja_servicio(
                    nombre=limpiar_valor(row.get("Nombre")),
                    direccion=limpiar_valor(row.get("Dirección")),
                    telefono=limpiar_valor(row.get("Tel")),
                    fecha=fecha_row.strftime("%d/%m/%Y") if fecha_row else "",
                    hora=rango_row,
                    folio=folio_ev,
                    origen=limpiar_valor(row.get("Origen")),
                    items=items_ev,
                    subtotal=(subtotal_ev or 0),
                    descuento=monto_desc_ev,
                    descuento_pct=desc_pct_ev,
                    iva=iva_ev,
                    total=(total_ev or 0),
                    ciudad=USUARIOS[st.session_state["usuario"]].get("ciudad", ""),
                    template_path=USUARIOS[st.session_state["usuario"]].get("template_pdf") or "assets/Hoja de servicio de Maxi Clean.pdf"
                )
                st.download_button(
                    "📄 Descargar hoja de servicio",
                    data=pdf_bytes_ev,
                    file_name=f"hoja_{limpiar_valor(row.get('Nombre','cliente')).replace(' ','_')}.pdf",
                    mime="application/pdf",
                    key=f"download_hoja_{id_click}"
                )
            except Exception as e_pdf:
                st.warning(f"No se pudo generar la hoja: {e_pdf}")

            if not realizado:
                if st.button("✅ Marcar como realizado y guardar en Sheets", use_container_width=True, type="primary"):
                    try:
                        client_auth = get_supabase_auth()
                        client_auth.table("clientes").update({"realizado": True}).eq("id", id_click).execute()

                        año_row = int(row.get("Año", row["Fecha"].year))
                        if año_row in SHEET_IDS:
                            client_gs = get_gspread_client()
                            sh = client_gs.open_by_key(SHEET_IDS[año_row])
                            worksheet = sh.get_worksheet(0)

                            col_b = worksheet.col_values(2)
                            ultimo_folio = 0
                            for v in col_b[1:]:
                                try:
                                    num = int(str(v).strip().split("/")[0])
                                    if num < 10000:
                                        ultimo_folio = max(ultimo_folio, num)
                                except:
                                    continue

                            siguiente_folio = ultimo_folio + 1
                            folio_interno = f"{siguiente_folio}/{str(año_row)[-2:]}"
                            origen_real = limpiar_valor(row.get("Origen"), "Int")
                            if origen_real.lower() in ["agenda", ""]:
                                origen_real = "Int"

                            id_cli = row.get("ID Cliente")
                            id_cli_limpio = int(id_cli) if id_cli and str(id_cli) not in ["", "nan", "None"] and pd.notnull(id_cli) else ""
                            es_rep = origen_real.lower() == "rep"
                            comentario_check = limpiar_valor(row.get("Comentarios con llamada posterior a venta"), "").lower()
                            cliente_no_contactar = any(p in comentario_check for p in [
                                "no se llama", "no contactar", "no vuelve", "cliente no deseable", "muy mal"
                            ])
                            serv_sheets = limpiar_valor(row.get("Servicio"))

                            nueva_fila = construir_fila_sheet(
                                siguiente_folio=siguiente_folio,
                                folio_interno=folio_interno,
                                fecha_str=row["Fecha"].strftime("%m/%d/%Y"),
                                id_cli_limpio=id_cli_limpio,
                                nombre=limpiar_valor(row.get("Nombre")),
                                tel=limpiar_valor(row.get("Tel")),
                                direccion=limpiar_valor(row.get("Dirección")),
                                origen=origen_real,
                                monto=float(row.get("Monto", 0)) if pd.notnull(row.get("Monto", 0)) else 0,
                                servicio=serv_sheets,
                                comentarios=""
                            )

                            col_c = worksheet.col_values(3)
                            datos_col_c = col_c[1:]
                            primera_vacia = None
                            for i, v in enumerate(datos_col_c):
                                if str(v).strip() == "":
                                    primera_vacia = i + 2
                                    break
                            if primera_vacia is None:
                                primera_vacia = len(datos_col_c) + 2

                            worksheet.insert_row(nueva_fila, primera_vacia)
                            client_auth.table("clientes").update({"folio": folio_interno}).eq("id", id_click).execute()

                            if cliente_no_contactar:
                                worksheet.format(f"A{primera_vacia}", {"backgroundColor": {"red": 0.9, "green": 0.2, "blue": 0.2}})
                            elif es_rep:
                                worksheet.format(f"A{primera_vacia}", {"backgroundColor": {"red": 0.2, "green": 0.8, "blue": 0.2}})

                            st.success("✅ Servicio marcado como realizado y guardado en Sheets.")
                        else:
                            st.success("✅ Marcado como realizado.")

                        st.cache_data.clear()
                        del st.session_state["evento_seleccionado"]
                        st.session_state["agenda_refresh"] = st.session_state.get("agenda_refresh", 0) + 1
                        import time as t
                        t.sleep(1)
                        st.rerun()

                    except Exception as e:
                        st.error(f"Error: {e}")
            else:
                st.success("✅ Este servicio ya fue realizado y guardado en Sheets.")

            # 💰 Sección de pago
            if features.get("cobranza"):
                st.markdown("#### 💰 Pago")
                estado_actual = limpiar_valor(row.get("estado_pago"), "pendiente") or "pendiente"
                forma_actual = limpiar_valor(row.get("forma_pago"), "")
                opciones_estado = ["pendiente", "pagado", "parcial"]
                opciones_forma = ["", "Efectivo", "Transferencia", "Tarjeta", "Cheque"]
                monto_serv = float(row.get("Monto", 0)) if pd.notnull(row.get("Monto", 0)) else 0.0
                with st.form("pago_form"):
                    nuevo_estado = st.selectbox("Estado de pago", opciones_estado,
                        index=opciones_estado.index(estado_actual) if estado_actual in opciones_estado else 0)
                    nueva_forma = st.selectbox("Forma de pago", opciones_forma,
                        index=opciones_forma.index(forma_actual) if forma_actual in opciones_forma else 0)
                    nuevo_monto_pagado = st.number_input("Monto pagado (para parcial)", min_value=0.0,
                        value=float(row.get("monto_pagado", 0) or 0), step=50.0)
                    guardar_pago = st.form_submit_button("💾 Guardar pago", use_container_width=True)
                    if guardar_pago:
                        try:
                            client_auth = get_supabase_auth()
                            upd = {"estado_pago": nuevo_estado, "forma_pago": nueva_forma or None}
                            if nuevo_estado == "pagado":
                                upd["fecha_pago"] = datetime.now().date().isoformat()
                                upd["monto_pagado"] = monto_serv
                            elif nuevo_estado == "parcial":
                                upd["monto_pagado"] = float(nuevo_monto_pagado)
                            else:
                                upd["monto_pagado"] = 0
                                upd["fecha_pago"] = None
                            client_auth.table("clientes").update(upd).eq("id", id_click).execute()
                            st.success("✅ Pago actualizado.")
                            st.cache_data.clear()
                            del st.session_state["evento_seleccionado"]
                            st.session_state["agenda_refresh"] = st.session_state.get("agenda_refresh", 0) + 1
                            import time as t
                            t.sleep(1)
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error al guardar pago: {e}")
            # 🧾 Facturación
            if features.get("comercial"):
                st.markdown("#### 🧾 Facturación")
                tipo_actual = limpiar_valor(row.get("tipo_cliente"), "Residencial") or "Residencial"
                with st.form("factura_form"):
                    nuevo_tipo = st.selectbox("Tipo de cliente", ["Residencial", "Comercial"],
                        index=1 if tipo_actual == "Comercial" else 0)
                    nueva_razon = st.text_input("Razón social", value=limpiar_valor(row.get("razon_social")))
                    nuevo_rfc = st.text_input("RFC", value=limpiar_valor(row.get("rfc")))
                    nueva_factura = st.text_input("Número de factura", value=limpiar_valor(row.get("factura")), placeholder="Ej. FAC-2026-0932")
                    guardar_fact = st.form_submit_button("💾 Guardar facturación", use_container_width=True)
                    if guardar_fact:
                        try:
                            client_auth = get_supabase_auth()
                            client_auth.table("clientes").update({
                                "tipo_cliente": nuevo_tipo,
                                "razon_social": nueva_razon or None,
                                "rfc": nuevo_rfc or None,
                                "factura": nueva_factura or None
                            }).eq("id", id_click).execute()
                            st.success("✅ Datos de facturación guardados.")
                            st.cache_data.clear()
                            del st.session_state["evento_seleccionado"]
                            st.session_state["agenda_refresh"] = st.session_state.get("agenda_refresh", 0) + 1
                            import time as t
                            t.sleep(1)
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")
            st.markdown("#### ✏️ Editar datos del servicio")
            with st.form("editar_servicio_form"):
                edit_nombre = st.text_input("Nombre", value=limpiar_valor(row.get("Nombre")))
                edit_tel = st.text_input("Teléfono", value=limpiar_valor(row.get("Tel")))
                edit_dir = st.text_input("Dirección", value=limpiar_valor(row.get("Dirección")))
                edit_servicio = st.text_input("Servicio", value=limpiar_valor(row.get("Servicio")))
                monto_actual = float(row.get("Monto", 0)) if pd.notnull(row.get("Monto", 0)) else 0.0
                edit_monto_str = st.text_input("Monto", value=f"{monto_actual:g}")
                try:
                    edit_monto = float(str(edit_monto_str).replace(",", "").replace("$", "").strip())
                except:
                    edit_monto = monto_actual
                origen_actual = limpiar_valor(row.get("Origen"), ORIGENES[0] if ORIGENES else "")
                edit_origen = st.selectbox(
                    "Origen", ORIGENES,
                    index=ORIGENES.index(origen_actual) if origen_actual in ORIGENES else 0
                )
                guardar_edicion = st.form_submit_button("💾 Guardar cambios", use_container_width=True)

                if guardar_edicion:
                    try:
                        client_auth = get_supabase_auth()
                        client_auth.table("clientes").update({
                            "nombre": edit_nombre,
                            "tel": edit_tel,
                            "direccion": edit_dir,
                            "servicio": edit_servicio,
                            "monto": float(edit_monto),
                            "origen": edit_origen
                        }).eq("id", id_click).execute()

                        folio_actual = str(limpiar_valor(row.get("folio")) or "").strip()
                        if realizado and SHEET_IDS:
                            try:
                                año_row = int(row.get("Año", row["Fecha"].year))
                                if año_row in SHEET_IDS:
                                    client_gs = get_gspread_client()
                                    sh = client_gs.open_by_key(SHEET_IDS[año_row])
                                    worksheet = sh.get_worksheet(0)

                                    def _norm_fecha(s):
                                        for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%y", "%d/%m/%y"):
                                            try:
                                                return datetime.strptime(str(s).strip(), fmt).date()
                                            except:
                                                continue
                                        return None

                                    fila_sheet = None
                                    if folio_actual:
                                        col_folio_idx = SHEET_COLS.get("folio_interno", 1) + 1
                                        for i, val in enumerate(worksheet.col_values(col_folio_idx)):
                                            if str(val).strip() == folio_actual:
                                                fila_sheet = i + 1
                                                break
                                    if fila_sheet is None:
                                        col_nombre = worksheet.col_values(SHEET_COLS.get("nombre", 4) + 1)
                                        col_fecha = worksheet.col_values(SHEET_COLS.get("fecha", 2) + 1)
                                        nombre_buscar = limpiar_valor(row.get("Nombre")).strip().lower()
                                        fecha_target = row["Fecha"].date() if pd.notnull(row["Fecha"]) else None
                                        n_filas = max(len(col_nombre), len(col_fecha))
                                        for i in range(n_filas):
                                            nv = col_nombre[i].strip().lower() if i < len(col_nombre) else ""
                                            fv = col_fecha[i] if i < len(col_fecha) else ""
                                            if nv == nombre_buscar and _norm_fecha(fv) == fecha_target:
                                                fila_sheet = i + 1
                                                break

                                    if fila_sheet:
                                        n = SHEET_COLS.get("nombre", 4) + 1
                                        t_ = SHEET_COLS.get("tel", 5) + 1
                                        d = SHEET_COLS.get("direccion", 6) + 1
                                        o = SHEET_COLS.get("origen", 7) + 1
                                        m = SHEET_COLS.get("monto", 8) + 1
                                        s = SHEET_COLS.get("servicio", 9) + 1
                                        from openpyxl.utils import get_column_letter
                                        worksheet.update(f"{get_column_letter(n)}{fila_sheet}", [[edit_nombre]])
                                        worksheet.update(f"{get_column_letter(t_)}{fila_sheet}", [[edit_tel]])
                                        worksheet.update(f"{get_column_letter(d)}{fila_sheet}", [[edit_dir]])
                                        worksheet.update(f"{get_column_letter(o)}{fila_sheet}", [[edit_origen]])
                                        worksheet.update(f"{get_column_letter(m)}{fila_sheet}", [[float(edit_monto)]])
                                        worksheet.update(f"{get_column_letter(s)}{fila_sheet}", [[edit_servicio]])
                                        st.success(f"✅ Actualizado en Supabase y Sheets (fila {fila_sheet}).")
                                    else:
                                        st.warning(f"Actualizado en Supabase, pero no encontré la fila en el Sheet {año_row}. folio='{folio_actual}', nombre='{limpiar_valor(row.get('Nombre'))}', fecha='{row['Fecha'].strftime('%m/%d/%Y') if pd.notnull(row['Fecha']) else ''}'.")
                                else:
                                    st.warning(f"Actualizado en Supabase, pero el año {año_row} no está en SHEET_IDS ({list(SHEET_IDS.keys())}).")
                            except Exception as e_sheet:
                                st.warning(f"Actualizado en Supabase pero error en Sheets: {e_sheet}")
                        else:
                            st.warning(f"Solo se actualizó en Supabase. Motivo → realizado={realizado}, sheets={bool(SHEET_IDS)}.")

                        st.cache_data.clear()
                        del st.session_state["evento_seleccionado"]
                        st.session_state["agenda_refresh"] = st.session_state.get("agenda_refresh", 0) + 1
                        import time as t
                        t.sleep(1)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al editar: {e}")

            st.markdown("#### 📅 Corregir fecha y hora")
            with st.form("editar_fecha_form"):
                col1, col2 = st.columns(2)
                with col1:
                    nueva_fecha = st.date_input("Nueva fecha:", value=row["Fecha"].date())
                with col2:
                    nueva_hora = st.time_input("Nueva hora:", value=None)
                confirmar_edicion = st.form_submit_button("📅 Cambiar fecha/hora", use_container_width=True)

                if confirmar_edicion:
                    try:
                        client_auth = get_supabase_auth()
                        nueva_hora_str = nueva_hora.strftime("%H:%M") if nueva_hora else None
                        client_auth.table("clientes").update({
                            "fecha": nueva_fecha.isoformat(),
                            "año": nueva_fecha.year,
                            "hora": nueva_hora_str
                        }).eq("id", id_click).execute()

                        st.success(f"✅ Fecha cambiada a {nueva_fecha.strftime('%d/%m/%Y')}")
                        st.cache_data.clear()
                        del st.session_state["evento_seleccionado"]
                        st.session_state["agenda_refresh"] = st.session_state.get("agenda_refresh", 0) + 1
                        import time as t
                        t.sleep(1)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al cambiar fecha: {e}")

            st.markdown("#### 🗑️ Eliminar servicio")
            with st.form("borrar_form"):
                st.warning("Esta acción no se puede deshacer.")
                confirmar_borrado = st.checkbox("Confirmo que quiero eliminar este servicio")
                borrar_btn = st.form_submit_button("🗑️ Eliminar", use_container_width=True)

                if borrar_btn:
                    if not confirmar_borrado:
                        st.error("Marca la casilla de confirmación antes de eliminar.")
                    else:
                        try:
                            client_auth = get_supabase_auth()
                            client_auth.table("clientes").delete().eq("id", id_click).execute()
                            st.success("✅ Servicio eliminado correctamente.")
                            st.cache_data.clear()
                            del st.session_state["evento_seleccionado"]
                            st.session_state["agenda_refresh"] = st.session_state.get("agenda_refresh", 0) + 1
                            import time as t
                            t.sleep(1)
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error al eliminar: {e}")
        else:
            st.warning("No se encontró el servicio seleccionado.")
            if st.button("Limpiar selección"):
                del st.session_state["evento_seleccionado"]
                st.rerun()
    # ── COTIZACIONES ──
elif pagina == "Cotizaciones":
    import base64
    import streamlit.components.v1 as components

    st.title("Cotizador de Servicios")

    cotizador = USUARIOS[st.session_state["usuario"]].get("cotizador", {})

    PAQUETES = cotizador.get("paquetes", [])
    MINIMO = cotizador.get("minimo", 0)
    PRECIOS = cotizador.get("precios", {})
    SERVICIOS_CON_CANTIDAD = cotizador.get("servicios_cantidad", [])
    SERVICIOS_CON_PLAZAS = cotizador.get("servicios_plazas", [])
    SERVICIOS_CON_SILLAS = cotizador.get("servicios_sillas", [])
    INTRO = cotizador.get("intro", "")
    PURT_DESC = cotizador.get("purt_descripcion", "")
    PURT_COSTO = cotizador.get("purt_costo", 0)
    DESC_PAQUETES = cotizador.get("descripcion_paquetes", {})
    DESCUENTOS = cotizador.get("descuentos_paquete", {})
    CIERRE = cotizador.get("cierre", "")
    FIRMA = cotizador.get("firma", "")

    if not PAQUETES or not PRECIOS:
        st.info("Este usuario no tiene cotizador configurado.")
        st.stop()

    st.markdown("### Agregar servicios a la cotización")

    if "items_cotizacion" not in st.session_state:
        st.session_state["items_cotizacion"] = []

    col1, col2 = st.columns([3, 1])

    with col1:
        servicio = st.selectbox("Servicio:", list(PRECIOS.keys()))

    with col2:
        if servicio in SERVICIOS_CON_CANTIDAD:
            cantidad = st.number_input("m2:", min_value=1, value=1)
            label_cantidad = "m2"
        elif servicio in SERVICIOS_CON_PLAZAS:
            cantidad = st.number_input("Plazas:", min_value=1, value=1)
            label_cantidad = "plazas"
        elif servicio in SERVICIOS_CON_SILLAS:
            cantidad = st.number_input("Sillas:", min_value=1, value=1)
            label_cantidad = "sillas"
        else:
            cantidad = 1
            label_cantidad = "unidad"

    precio_data = {
        "Paquete": PAQUETES,
        "Precio": [f"${PRECIOS[servicio][p] * cantidad:,.0f}" for p in PAQUETES]
    }
    st.dataframe(pd.DataFrame(precio_data), use_container_width=True, hide_index=True)

    if st.button("Agregar a cotización", use_container_width=True):
        st.session_state["items_cotizacion"].append({
            "Servicio": servicio,
            "Cantidad": cantidad,
            "Label": label_cantidad,
            "Precios": {p: PRECIOS[servicio][p] * cantidad for p in PAQUETES}
        })

    if st.session_state["items_cotizacion"]:
        st.markdown("---")
        st.markdown("### Resumen — todos los paquetes")

        filas = []
        for item in st.session_state["items_cotizacion"]:
            fila = {"Servicio": f"{item['Servicio']} ({item['Cantidad']} {item['Label']})"}
            for p in PAQUETES:
                fila[p] = f"${item['Precios'][p]:,.0f}"
            filas.append(fila)

        totales = {"Servicio": "TOTAL"}
        for p in PAQUETES:
            total_p = sum(i["Precios"][p] for i in st.session_state["items_cotizacion"])
            descuento = DESCUENTOS.get(p, 0)
            total_con_descuento = total_p * (1 - descuento / 100)
            total_final = max(total_con_descuento, MINIMO)
            nota = " *" if total_p < MINIMO else ""
            totales[p] = f"${total_final:,.0f}{nota}"

        filas.append(totales)
        st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

        # ── MENSAJE PROFESIONAL ──
        st.markdown("### Mensaje para cliente")

        nombre_cliente = st.text_input("Nombre del cliente (opcional):")
        incluir_purt = st.checkbox("Incluir descripción de PURT", value=bool(PURT_DESC))

        lineas = []

        if nombre_cliente:
            lineas.append(f"Estimado/a {nombre_cliente},\n")

        if INTRO:
            lineas.append(INTRO)
            lineas.append("")

        if incluir_purt and PURT_DESC:
            lineas.append(PURT_DESC)
            lineas.append("")

        lineas.append("El servicio de acuerdo a su solicitud tiene una inversión de:\n")

        for p in PAQUETES:
            total_p = sum(i["Precios"][p] for i in st.session_state["items_cotizacion"])
            descuento = DESCUENTOS.get(p, 0)
            total_con_descuento = total_p * (1 - descuento / 100)
            total_final = max(total_con_descuento, MINIMO)

            lineas.append(f"Paquete {p}:")
            if p in DESC_PAQUETES:
                lineas.append(DESC_PAQUETES[p])

            for item in st.session_state["items_cotizacion"]:
                lineas.append(f"{item['Servicio']} ({item['Cantidad']} {item['Label']}): ${item['Precios'][p]:,.0f}")

            if incluir_purt and PURT_COSTO > 0:
                lineas.append(f"PURT: ${PURT_COSTO:,.0f}")

            if descuento > 0:
                lineas.append(f"Descuento por volumen {descuento}%")

            lineas.append(f"TOTAL: ${total_final:,.0f}")
            lineas.append("")

        if CIERRE:
            lineas.append(CIERRE)

        if FIRMA:
            lineas.append(FIRMA)

        st.text_area("Copia este mensaje:", "\n".join(lineas), height=400)

        # ── PDF ──
        st.markdown("### 📄 Descargar cotización")

        nombre_pdf = nombre_cliente if nombre_cliente else "Cliente"
        fecha_pdf = datetime.now().strftime("%d/%m/%Y")
        empresa_pdf = st.session_state.get("empresa", "")

        filas_pdf = ""
        for p in PAQUETES:
            total_p = sum(i["Precios"][p] for i in st.session_state["items_cotizacion"])
            descuento = DESCUENTOS.get(p, 0)
            total_con_descuento = total_p * (1 - descuento / 100)
            total_final = max(total_con_descuento, MINIMO)

            servicios_html = ""
            for item in st.session_state["items_cotizacion"]:
                servicios_html += f"""
                <tr>
                    <td style='padding:6px 12px; border-bottom:1px solid #eee;'>
                        {item['Servicio']} ({item['Cantidad']} {item['Label']})
                    </td>
                    <td style='padding:6px 12px; text-align:right; border-bottom:1px solid #eee;'>
                        ${item['Precios'][p]:,.0f}
                    </td>
                </tr>
                """

            if incluir_purt and PURT_COSTO > 0:
                servicios_html += f"""
                <tr>
                    <td style='padding:6px 12px; border-bottom:1px solid #eee;'>PURT</td>
                    <td style='padding:6px 12px; text-align:right; border-bottom:1px solid #eee;'>
                        ${PURT_COSTO:,.0f}
                    </td>
                </tr>
                """

            if descuento > 0:
                servicios_html += f"""
                <tr>
                    <td style='padding:6px 12px; border-bottom:1px solid #eee; color:#888;'>
                        Descuento {descuento}%
                    </td>
                    <td style='padding:6px 12px; text-align:right; border-bottom:1px solid #eee; color:#888;'>
                        -${total_p * descuento / 100:,.0f}
                    </td>
                </tr>
                """

            desc_paquete_html = f"<p style='color:#555; font-size:13px; margin:4px 0 8px 0;'>{DESC_PAQUETES[p]}</p>" if p in DESC_PAQUETES else ""

            filas_pdf += f"""
            <div style="margin-bottom:24px; border:1px solid #ddd; border-radius:8px; overflow:hidden;">
                <div style="background:#2B5BAA; color:white; padding:10px 16px; font-weight:bold; font-size:15px;">
                    Paquete {p}
                </div>
                <div style="padding:8px 12px;">
                    {desc_paquete_html}
                </div>
                <table style="width:100%; border-collapse:collapse; font-size:14px;">
                    {servicios_html}
                    <tr style="background:#f0f4ff;">
                        <td style="padding:10px 12px; font-weight:bold;">TOTAL</td>
                        <td style="padding:10px 12px; text-align:right; font-weight:bold; font-size:16px;">
                            ${total_final:,.0f}
                        </td>
                    </tr>
                </table>
            </div>
            """

        intro_html = f"<p style='margin-bottom:16px;'>{INTRO}</p>" if INTRO else ""
        purt_html = f"<p style='margin-bottom:16px; color:#555;'>{PURT_DESC}</p>" if (incluir_purt and PURT_DESC) else ""
        cierre_html = f"<p style='margin-top:8px;'>{CIERRE}</p>" if CIERRE else ""
        firma_html = f"<p style='margin-top:4px; color:#555;'>{FIRMA}</p>" if FIRMA else ""

        html_pdf = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Cotización — {nombre_pdf}</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    padding: 48px;
                    max-width: 680px;
                    margin: auto;
                    color: #333;
                    font-size: 14px;
                }}
                h1 {{ color: #2B5BAA; margin-bottom: 4px; font-size: 28px; }}
                .meta {{ color: #666; font-size: 13px; margin-bottom: 24px; line-height: 1.8; }}
                .divider {{ border: none; border-top: 1px solid #eee; margin: 24px 0; }}
                .footer {{ margin-top: 40px; font-size: 13px; color: #888;
                           border-top: 1px solid #eee; padding-top: 16px; }}
                @media print {{
                    body {{ padding: 24px; }}
                }}
            </style>
        </head>
        <body>
            <h1>Cotización de servicios</h1>
            <div class="meta">
                <strong>Para:</strong> {nombre_pdf}<br>
                <strong>De:</strong> {empresa_pdf}<br>
                <strong>Fecha:</strong> {fecha_pdf}
            </div>

            <hr class="divider">

            {intro_html}
            {purt_html}

            <p style="margin-bottom:20px;">
                El servicio de acuerdo a su solicitud tiene una inversión de:
            </p>

            {filas_pdf}

            <div class="footer">
                {cierre_html}
                {firma_html}
            </div>
        </body>
        </html>
        """

        b64_pdf = base64.b64encode(html_pdf.encode("utf-8")).decode("utf-8")
        nombre_archivo = f"cotizacion_{nombre_pdf.replace(' ', '_')}_{datetime.now().strftime('%d%m%Y')}.html"

        components.html(f"""
        <a href="data:text/html;base64,{b64_pdf}"
           download="{nombre_archivo}"
           style="
               display:block;
               background-color:#2B5BAA;
               color:white;
               text-decoration:none;
               text-align:center;
               border-radius:8px;
               padding:14px 24px;
               font-size:16px;
               font-family:Arial, sans-serif;
               margin-top:8px;
           ">
            📄 Descargar cotización
        </a>
        """, height=65)

        st.caption("Se descarga como archivo HTML. Ábrelo en el navegador y usa Cmd+P / Ctrl+P → Guardar como PDF.")

        st.markdown("---")

        if st.button("Limpiar cotización", use_container_width=True):
            st.session_state["items_cotizacion"] = []
            st.rerun()
elif pagina == "Mi cuenta":
    st.title("🔑 Mi cuenta")
    st.write(f"**Negocio:** {st.session_state.get('empresa', '')}")
    st.markdown("### Cambiar contraseña")
    with st.form("cambiar_pass_form"):
        nueva_pass = st.text_input("Nueva contraseña", type="password")
        confirmar_pass = st.text_input("Confirmar contraseña", type="password")
        if st.form_submit_button("Actualizar contraseña"):
            if len(nueva_pass) < 6:
                st.error("La contraseña debe tener al menos 6 caracteres.")
            elif nueva_pass != confirmar_pass:
                st.error("Las contraseñas no coinciden.")
            else:
                try:
                    sb = create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])
                    sb.auth.set_session(st.session_state["access_token"], st.session_state.get("refresh_token", ""))
                    sb.auth.update_user({"password": nueva_pass})
                    st.success("✅ Contraseña actualizada. Úsala la próxima vez que inicies sesión.")
                except Exception as e:
                    st.error(f"Error al cambiar contraseña: {e}")